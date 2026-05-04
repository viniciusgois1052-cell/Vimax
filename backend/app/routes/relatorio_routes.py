from flask import Blueprint, jsonify, request, send_file
from sqlalchemy import func, extract, text
from .. import db
from ..models.chamado import Chamado
from ..models.ativo import Ativo
from ..models.empresa import Empresa
from ..models.orcamento import Orcamento
from ..models.contrato import Contrato
from ..models.categoria_chamado import CategoriaChamado
from ..models.localizacao import Localizacao
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

relatorio_bp = Blueprint("relatorio_bp", __name__)

# ── helpers ──────────────────────────────────────────────────────────────────

def _style_header(ws, num_cols):
    """Aplica estilo de cabeçalho na primeira linha da planilha."""
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20
    for i in range(1, num_cols + 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 22

def _make_excel(sheets):
    """
    sheets: lista de dicts { title, headers, rows }
    Retorna BytesIO com o arquivo Excel.
    """
    wb = Workbook()
    first = True
    for sheet in sheets:
        if first:
            ws = wb.active
            ws.title = sheet["title"][:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet["title"][:31])
        ws.append(sheet["headers"])
        for row in sheet["rows"]:
            ws.append(row)
        _style_header(ws, len(sheet["headers"]))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _get_token(request):
    return request.headers.get("X-API-Token") or request.args.get("token")

# ── DASHBOARD ────────────────────────────────────────────────────────────────

@relatorio_bp.route("/dashboard", methods=["GET"])
def dashboard_relatorios():
    try:
        # Filtros de data
        data_inicio = request.args.get('data_inicio')
        data_fim    = request.args.get('data_fim')

        def filtro_data_orm(query, campo):
            if data_inicio:
                query = query.filter(campo >= data_inicio)
            if data_fim:
                query = query.filter(campo <= data_fim + " 23:59:59")
            return query

        def filtro_data_sql():
            clauses = []
            if data_inicio:
                clauses.append(f"AND c.created_at >= '{data_inicio}'")
            if data_fim:
                clauses.append(f"AND c.created_at <= '{data_fim} 23:59:59'")
            return " ".join(clauses)

        filtro_sql = filtro_data_sql()

        # Estatísticas por Empresa
        empresas = Empresa.query.all()
        stats_map = {}
        for emp in empresas:
            stats_map[emp.id] = {
                "nome": emp.nome,
                "gasto_orcamentos": 0.0,
                "custo_chamados": 0.0,
                "ativos": 0,
                "ativos_sem_contrato": 0
            }

        for emp_id, total in (db.session.query(Orcamento.empresa_id, func.sum(Orcamento.valor))
                               .filter(Orcamento.status == 'Aprovado')
                               .group_by(Orcamento.empresa_id).all()):
            if emp_id in stats_map:
                stats_map[emp_id]["gasto_orcamentos"] = float(total or 0)

        q_custo = db.session.query(Chamado.empresa_id, func.sum(Chamado.valor_total)).filter(Chamado.ativo == True)
        q_custo = filtro_data_orm(q_custo, Chamado.created_at)
        for emp_id, total in q_custo.group_by(Chamado.empresa_id).all():
            if emp_id in stats_map:
                stats_map[emp_id]["custo_chamados"] = float(total or 0)

        for emp_id, qtd in db.session.query(Ativo.empresa_id, func.count(Ativo.id)).group_by(Ativo.empresa_id).all():
            if emp_id in stats_map:
                stats_map[emp_id]["ativos"] = int(qtd or 0)

        for emp_id, qtd in (db.session.query(Ativo.empresa_id, func.count(Ativo.id))
                             .filter(Ativo.contrato_id == None)
                             .group_by(Ativo.empresa_id).all()):
            if emp_id in stats_map:
                stats_map[emp_id]["ativos_sem_contrato"] = int(qtd or 0)

        empresas_stats = sorted(list(stats_map.values()), key=lambda x: x["nome"])

        # Custos chamados por ativo (top 10)
        custos_ativos_chamados = []
        try:
            q = (db.session.query(Ativo.nome, func.sum(Chamado.valor_total))
                    .join(Chamado, Chamado.ativo_id == Ativo.id)
                    .filter(Chamado.ativo == True))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Ativo.nome).order_by(func.sum(Chamado.valor_total).desc()).limit(10).all()
            custos_ativos_chamados = [[r[0], float(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro custos_ativos_chamados: {e}")

        # Gastos orçamentos por ativo (top 10)
        gastos_ativos_orcamentos = []
        try:
            rows = (db.session.query(Ativo.nome, func.sum(Orcamento.valor))
                    .join(Orcamento, Ativo.orcamento_id == Orcamento.id)
                    .group_by(Ativo.nome)
                    .order_by(func.sum(Orcamento.valor).desc())
                    .limit(10).all())
            gastos_ativos_orcamentos = [[r[0], float(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro gastos_ativos_orcamentos: {e}")

        # Ativos sem contrato
        ativos_sem_contrato_count = (db.session.query(func.count(Ativo.id))
                                     .filter(Ativo.contrato_id == None).scalar() or 0)

        ativos_sem_contrato_detalhes = []
        try:
            rows = (db.session.query(Ativo.nome, Empresa.nome, Localizacao.nome)
                    .join(Empresa, Ativo.empresa_id == Empresa.id)
                    .outerjoin(Localizacao, Ativo.localizacao_id == Localizacao.id)
                    .filter(Ativo.contrato_id == None)
                    .order_by(Empresa.nome, Ativo.nome).all())
            ativos_sem_contrato_detalhes = [[r[0], r[1], r[2] or 'Não informada'] for r in rows]
        except Exception as e:
            print(f"Erro ativos_sem_contrato_detalhes: {e}")

        # Chamados por categoria
        chamados_por_categoria = []
        try:
            q = (db.session.query(CategoriaChamado.nome, func.count(Chamado.id))
                    .join(Chamado, Chamado.categoria_id == CategoriaChamado.id)
                    .filter(Chamado.ativo == True))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(CategoriaChamado.nome).all()
            chamados_por_categoria = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_categoria: {e}")

        # Chamados por status
        chamados_por_status = []
        try:
            q = db.session.query(Chamado.status, func.count(Chamado.id)).filter(Chamado.ativo == True)
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Chamado.status).all()
            chamados_por_status = [[r[0] or 'Sem status', int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_status: {e}")

        # Chamados por tipo (maquinario / infraestrutura)
        chamados_por_tipo = []
        try:
            q = db.session.query(Chamado.tipo, func.count(Chamado.id)).filter(Chamado.ativo == True)
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Chamado.tipo).all()
            chamados_por_tipo = [[r[0] or 'Não informado', int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_tipo: {e}")

        # Chamados por prioridade
        chamados_por_prioridade = []
        try:
            q = db.session.query(Chamado.prioridade, func.count(Chamado.id)).filter(Chamado.ativo == True)
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Chamado.prioridade).all()
            chamados_por_prioridade = [[r[0] or 'Não informada', int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_prioridade: {e}")

        # Chamados por empresa
        chamados_por_empresa = []
        try:
            q = (db.session.query(Empresa.nome, func.count(Chamado.id))
                    .join(Chamado, Chamado.empresa_id == Empresa.id)
                    .filter(Chamado.ativo == True))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Empresa.nome).order_by(func.count(Chamado.id).desc()).all()
            chamados_por_empresa = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_empresa: {e}")

        # Chamados por maquinário (ativo)
        chamados_por_maquinario = []
        try:
            q = (db.session.query(Ativo.nome, func.count(Chamado.id))
                    .join(Chamado, Chamado.ativo_id == Ativo.id)
                    .filter(Chamado.ativo == True, Chamado.tipo == 'maquinario'))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Ativo.nome).order_by(func.count(Chamado.id).desc()).limit(15).all()
            chamados_por_maquinario = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_maquinario: {e}")

        # Chamados por infraestrutura
        chamados_por_infraestrutura = []
        try:
            rows = db.session.execute(text(f"""
                SELECT i.nome, COUNT(c.id) as total
                FROM chamados c
                JOIN infraestrutura i ON c.infraestrutura_id = i.id
                WHERE c.ativo = 1 AND c.tipo = 'infraestrutura'
                {filtro_sql}
                GROUP BY i.nome
                ORDER BY total DESC
                LIMIT 15
            """)).fetchall()
            chamados_por_infraestrutura = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_infraestrutura: {e}")

        # Tempo médio de solução (em dias)
        tempo_medio_solucao = []
        try:
            rows = db.session.execute(text(f"""
                SELECT e.nome,
                       ROUND(AVG(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0), 1) as media_dias
                FROM chamados c
                JOIN empresas e ON c.empresa_id = e.id
                WHERE c.ativo = 1 AND c.data_solucao IS NOT NULL
                {filtro_sql}
                GROUP BY e.nome
                ORDER BY media_dias DESC
            """)).fetchall()
            tempo_medio_solucao = [[r[0], float(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro tempo_medio_solucao: {e}")

        # Evolução mensal
        chamados_por_mes = []
        try:
            q_mes = (db.session.query(
                        extract('month', Chamado.created_at),
                        extract('year', Chamado.created_at),
                        func.count(Chamado.id))
                    .filter(Chamado.ativo == True))
            q_mes = filtro_data_orm(q_mes, Chamado.created_at)
            rows = (q_mes
                    .group_by(extract('year', Chamado.created_at), extract('month', Chamado.created_at))
                    .order_by(extract('year', Chamado.created_at).desc(), extract('month', Chamado.created_at).desc())
                    .limit(12).all())
            rows = list(reversed(rows))
            meses = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            for r in rows:
                try:
                    chamados_por_mes.append([f"{meses[int(r[0])]}/{str(int(r[1]))[-2:]}", int(r[2])])
                except:
                    continue
        except Exception as e:
            print(f"Erro chamados_por_mes: {e}")

        # Resumo geral
        total_ativos = db.session.query(func.count(Ativo.id)).scalar() or 0
        q_total_ch = db.session.query(func.count(Chamado.id)).filter(Chamado.ativo == True)
        q_total_ch = filtro_data_orm(q_total_ch, Chamado.created_at)
        total_chamados = q_total_ch.scalar() or 0
        total_gasto_orcamentos = (db.session.query(func.sum(Orcamento.valor))
                                  .filter(Orcamento.status == 'Aprovado').scalar() or 0)
        q_custo_total = db.session.query(func.sum(Chamado.valor_total)).filter(Chamado.ativo == True)
        q_custo_total = filtro_data_orm(q_custo_total, Chamado.created_at)
        total_custo_chamados = q_custo_total.scalar() or 0
        total_infraestruturas = db.session.execute(
            text("SELECT COUNT(*) FROM infraestrutura WHERE ativo = 1")).scalar() or 0

        return jsonify({
            "empresas_stats": empresas_stats,
            "custos_ativos_chamados": custos_ativos_chamados,
            "gastos_ativos_orcamentos": gastos_ativos_orcamentos,
            "ativos_sem_contrato": int(ativos_sem_contrato_count),
            "ativos_sem_contrato_detalhes": ativos_sem_contrato_detalhes,
            "chamados_por_categoria": chamados_por_categoria,
            "chamados_por_status": chamados_por_status,
            "chamados_por_tipo": chamados_por_tipo,
            "chamados_por_prioridade": chamados_por_prioridade,
            "chamados_por_empresa": chamados_por_empresa,
            "chamados_por_maquinario": chamados_por_maquinario,
            "chamados_por_infraestrutura": chamados_por_infraestrutura,
            "tempo_medio_solucao": tempo_medio_solucao,
            "chamados_por_mes": chamados_por_mes,
            "resumo": {
                "total_ativos": int(total_ativos),
                "total_chamados": int(total_chamados),
                "total_gasto_orcamentos": float(total_gasto_orcamentos or 0),
                "total_custo_chamados": float(total_custo_chamados or 0),
                "ativos_sem_contrato": int(ativos_sem_contrato_count),
                "total_infraestruturas": int(total_infraestruturas),
            }
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── EXPORTS ──────────────────────────────────────────────────────────────────

@relatorio_bp.route("/export/chamados_completo", methods=["GET"])
def export_chamados_completo():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.id,
                c.titulo,
                c.status,
                c.prioridade,
                c.tipo,
                c.data_abertura,
                c.data_solucao,
                ROUND(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0, 1) as dias_solucao,
                e.nome as empresa,
                cat.nome as categoria,
                a.nome as ativo,
                i.nome as infraestrutura,
                l.nome as localizacao,
                c.valor_total,
                c.descricao
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            LEFT JOIN ativos a ON c.ativo_id = a.id
            LEFT JOIN infraestrutura i ON c.infraestrutura_id = i.id
            LEFT JOIN localizacoes l ON c.localizacao_id = l.id
            WHERE c.ativo = 1
            ORDER BY c.data_abertura DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados Completo",
            "headers": ["ID", "Título", "Status", "Prioridade", "Tipo", "Data Abertura",
                        "Data Solução", "Dias p/ Solução", "Empresa", "Categoria",
                        "Maquinário", "Infraestrutura", "Localização", "Custo Total", "Descrição"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_completo.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_maquinario", methods=["GET"])
def export_chamados_por_maquinario():
    try:
        rows = db.session.execute(text("""
            SELECT
                a.nome as maquinario,
                e.nome as empresa,
                COUNT(c.id) as total_chamados,
                SUM(c.valor_total) as custo_total,
                AVG(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0) as media_dias_solucao
            FROM chamados c
            JOIN ativos a ON c.ativo_id = a.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.tipo = 'maquinario'
            GROUP BY a.nome, e.nome
            ORDER BY total_chamados DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Maquinário",
            "headers": ["Maquinário", "Empresa", "Total Chamados", "Custo Total", "Média Dias p/ Solução"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_maquinario.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_maquinario_categoria", methods=["GET"])
def export_chamados_por_maquinario_categoria():
    try:
        rows = db.session.execute(text("""
            SELECT
                a.nome as maquinario,
                cat.nome as categoria,
                e.nome as empresa,
                COUNT(c.id) as total_chamados,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            JOIN ativos a ON c.ativo_id = a.id
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.tipo = 'maquinario'
            GROUP BY a.nome, cat.nome, e.nome
            ORDER BY a.nome, total_chamados DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Maquinário x Categoria",
            "headers": ["Maquinário", "Categoria", "Empresa", "Total Chamados", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_maquinario_categoria.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_infraestrutura", methods=["GET"])
def export_chamados_por_infraestrutura():
    try:
        rows = db.session.execute(text("""
            SELECT
                i.nome as infraestrutura,
                e.nome as empresa,
                COUNT(c.id) as total_chamados,
                SUM(c.valor_total) as custo_total,
                AVG(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0) as media_dias_solucao
            FROM chamados c
            JOIN infraestrutura i ON c.infraestrutura_id = i.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.tipo = 'infraestrutura'
            GROUP BY i.nome, e.nome
            ORDER BY total_chamados DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Infraestrutura",
            "headers": ["Infraestrutura", "Empresa", "Total Chamados", "Custo Total", "Média Dias p/ Solução"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_infraestrutura.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_empresa", methods=["GET"])
def export_chamados_por_empresa():
    try:
        rows = db.session.execute(text("""
            SELECT
                e.nome as empresa,
                c.status,
                c.prioridade,
                c.tipo,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY e.nome, c.status, c.prioridade, c.tipo
            ORDER BY e.nome, total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Empresa",
            "headers": ["Empresa", "Status", "Prioridade", "Tipo", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_empresa.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_categoria", methods=["GET"])
def export_chamados_por_categoria():
    try:
        rows = db.session.execute(text("""
            SELECT
                cat.nome as categoria,
                e.nome as empresa,
                c.tipo,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY cat.nome, e.nome, c.tipo
            ORDER BY total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Categoria",
            "headers": ["Categoria", "Empresa", "Tipo", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_categoria.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_status", methods=["GET"])
def export_chamados_por_status():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.status,
                e.nome as empresa,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY c.status, e.nome
            ORDER BY total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Status",
            "headers": ["Status", "Empresa", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_status.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_prioridade", methods=["GET"])
def export_chamados_por_prioridade():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.prioridade,
                e.nome as empresa,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY c.prioridade, e.nome
            ORDER BY total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Prioridade",
            "headers": ["Prioridade", "Empresa", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_prioridade.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_mes", methods=["GET"])
def export_chamados_por_mes():
    try:
        rows = db.session.execute(text("""
            SELECT
                DATE_FORMAT(c.created_at, '%m/%Y') as mes_ano,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            WHERE c.ativo = 1
            GROUP BY DATE_FORMAT(c.created_at, '%m/%Y'), DATE_FORMAT(c.created_at, '%Y-%m')
            ORDER BY DATE_FORMAT(c.created_at, '%Y-%m') DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Evolução Mensal",
            "headers": ["Mês/Ano", "Total Chamados", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_mes.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/tempo_solucao", methods=["GET"])
def export_tempo_solucao():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.id,
                c.titulo,
                e.nome as empresa,
                c.status,
                c.prioridade,
                c.tipo,
                c.data_abertura,
                c.data_solucao,
                ROUND(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0, 1) as dias_solucao
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.data_solucao IS NOT NULL
            ORDER BY dias_solucao DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Tempo de Solução",
            "headers": ["ID", "Título", "Empresa", "Status", "Prioridade", "Tipo",
                        "Data Abertura", "Data Solução", "Dias p/ Solução"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="tempo_solucao.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/custos_ativos_chamados", methods=["GET"])
def export_custos_ativos_chamados():
    try:
        rows = (db.session.query(Ativo.nome, Empresa.nome, func.count(Chamado.id), func.sum(Chamado.valor_total))
                .join(Chamado, Chamado.ativo_id == Ativo.id)
                .join(Empresa, Ativo.empresa_id == Empresa.id)
                .filter(Chamado.ativo == True)
                .group_by(Ativo.nome, Empresa.nome)
                .order_by(func.sum(Chamado.valor_total).desc()).all())

        buf = _make_excel([{
            "title": "Custos Chamados por Ativo",
            "headers": ["Ativo", "Empresa", "Total Chamados", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="custos_ativos_chamados.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/gastos_ativos_orcamentos", methods=["GET"])
def export_gastos_ativos_orcamentos():
    try:
        rows = (db.session.query(Ativo.nome, func.sum(Orcamento.valor))
                .join(Orcamento, Ativo.orcamento_id == Orcamento.id)
                .group_by(Ativo.nome)
                .order_by(func.sum(Orcamento.valor).desc()).all())

        buf = _make_excel([{
            "title": "Gastos Orçamentos por Ativo",
            "headers": ["Ativo", "Gasto Total Orçamentos"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="gastos_ativos_orcamentos.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/ativos_sem_contrato", methods=["GET"])
def export_ativos_sem_contrato():
    try:
        rows = (db.session.query(Ativo.nome, Empresa.nome, Localizacao.nome)
                .join(Empresa, Ativo.empresa_id == Empresa.id)
                .outerjoin(Localizacao, Ativo.localizacao_id == Localizacao.id)
                .filter(Ativo.contrato_id == None)
                .order_by(Empresa.nome, Ativo.nome).all())

        buf = _make_excel([{
            "title": "Ativos sem Contrato",
            "headers": ["Ativo", "Empresa", "Localização"],
            "rows": [[r[0], r[1], r[2] or 'Não informada'] for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="ativos_sem_contrato.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/empresas_stats", methods=["GET"])
def export_empresas_stats():
    try:
        empresas = Empresa.query.all()
        stats_map = {}
        for emp in empresas:
            stats_map[emp.id] = {"nome": emp.nome, "gasto_orcamentos": 0.0, "custo_chamados": 0.0, "ativos": 0, "ativos_sem_contrato": 0}

        for emp_id, total in (db.session.query(Orcamento.empresa_id, func.sum(Orcamento.valor))
                               .filter(Orcamento.status == 'Aprovado').group_by(Orcamento.empresa_id).all()):
            if emp_id in stats_map: stats_map[emp_id]["gasto_orcamentos"] = float(total or 0)

        for emp_id, total in (db.session.query(Chamado.empresa_id, func.sum(Chamado.valor_total))
                               .filter(Chamado.ativo == True).group_by(Chamado.empresa_id).all()):
            if emp_id in stats_map: stats_map[emp_id]["custo_chamados"] = float(total or 0)

        for emp_id, qtd in db.session.query(Ativo.empresa_id, func.count(Ativo.id)).group_by(Ativo.empresa_id).all():
            if emp_id in stats_map: stats_map[emp_id]["ativos"] = int(qtd or 0)

        for emp_id, qtd in (db.session.query(Ativo.empresa_id, func.count(Ativo.id))
                             .filter(Ativo.contrato_id == None).group_by(Ativo.empresa_id).all()):
            if emp_id in stats_map: stats_map[emp_id]["ativos_sem_contrato"] = int(qtd or 0)

        rows = sorted(list(stats_map.values()), key=lambda x: x["nome"])
        buf = _make_excel([{
            "title": "Estatísticas por Empresa",
            "headers": ["Empresa", "Qtd. Ativos", "Ativos sem Contrato", "Custo Chamados", "Gasto Orçamentos"],
            "rows": [[r["nome"], r["ativos"], r["ativos_sem_contrato"], r["custo_chamados"], r["gasto_orcamentos"]] for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="empresas_stats.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/tudo", methods=["GET"])
def export_tudo():
    """Exporta todas as planilhas em um único arquivo Excel com múltiplas abas."""
    try:
        sheets = []

        # Aba 1 - Chamados Completo
        rows1 = db.session.execute(text("""
            SELECT c.id, c.titulo, c.status, c.prioridade, c.tipo,
                   c.data_abertura, c.data_solucao,
                   ROUND(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao)/24.0,1),
                   e.nome, cat.nome, a.nome, i.nome, l.nome, c.valor_total
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            LEFT JOIN ativos a ON c.ativo_id = a.id
            LEFT JOIN infraestrutura i ON c.infraestrutura_id = i.id
            LEFT JOIN localizacoes l ON c.localizacao_id = l.id
            WHERE c.ativo = 1 ORDER BY c.data_abertura DESC
        """)).fetchall()
        sheets.append({"title": "Chamados Completo",
                        "headers": ["ID","Título","Status","Prioridade","Tipo","Abertura","Solução","Dias","Empresa","Categoria","Maquinário","Infraestrutura","Localização","Custo"],
                        "rows": [list(r) for r in rows1]})

        # Aba 2 - Por Maquinário
        rows2 = db.session.execute(text("""
            SELECT a.nome, e.nome, COUNT(c.id), SUM(c.valor_total),
                   ROUND(AVG(TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao)/24.0),1)
            FROM chamados c JOIN ativos a ON c.ativo_id=a.id
            JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 AND c.tipo='maquinario'
            GROUP BY a.nome,e.nome ORDER BY COUNT(c.id) DESC
        """)).fetchall()
        sheets.append({"title": "Por Maquinário",
                        "headers": ["Maquinário","Empresa","Total","Custo","Média Dias"],
                        "rows": [list(r) for r in rows2]})

        # Aba 3 - Por Infraestrutura
        rows3 = db.session.execute(text("""
            SELECT i.nome, e.nome, COUNT(c.id), SUM(c.valor_total),
                   ROUND(AVG(TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao)/24.0),1)
            FROM chamados c JOIN infraestrutura i ON c.infraestrutura_id=i.id
            JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 AND c.tipo='infraestrutura'
            GROUP BY i.nome,e.nome ORDER BY COUNT(c.id) DESC
        """)).fetchall()
        sheets.append({"title": "Por Infraestrutura",
                        "headers": ["Infraestrutura","Empresa","Total","Custo","Média Dias"],
                        "rows": [list(r) for r in rows3]})

        # Aba 4 - Por Empresa
        rows4 = db.session.execute(text("""
            SELECT e.nome, c.status, c.prioridade, c.tipo, COUNT(c.id), SUM(c.valor_total)
            FROM chamados c JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 GROUP BY e.nome,c.status,c.prioridade,c.tipo ORDER BY e.nome
        """)).fetchall()
        sheets.append({"title": "Por Empresa",
                        "headers": ["Empresa","Status","Prioridade","Tipo","Total","Custo"],
                        "rows": [list(r) for r in rows4]})

        # Aba 5 - Por Categoria
        rows5 = db.session.execute(text("""
            SELECT cat.nome, e.nome, c.tipo, COUNT(c.id), SUM(c.valor_total)
            FROM chamados c LEFT JOIN categorias_chamado cat ON c.categoria_id=cat.id
            JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 GROUP BY cat.nome,e.nome,c.tipo ORDER BY COUNT(c.id) DESC
        """)).fetchall()
        sheets.append({"title": "Por Categoria",
                        "headers": ["Categoria","Empresa","Tipo","Total","Custo"],
                        "rows": [list(r) for r in rows5]})

        # Aba 6 - Evolução Mensal
        rows6 = db.session.execute(text("""
            SELECT DATE_FORMAT(created_at,'%m/%Y'), COUNT(id), SUM(valor_total)
            FROM chamados WHERE ativo=1
            GROUP BY DATE_FORMAT(created_at,'%m/%Y'), DATE_FORMAT(created_at,'%Y-%m')
            ORDER BY DATE_FORMAT(created_at,'%Y-%m')
        """)).fetchall()
        sheets.append({"title": "Evolução Mensal",
                        "headers": ["Mês/Ano","Total","Custo Total"],
                        "rows": [list(r) for r in rows6]})

        # Aba 7 - Tempo de Solução
        rows7 = db.session.execute(text("""
            SELECT c.id, c.titulo, e.nome, c.prioridade, c.tipo,
                   c.data_abertura, c.data_solucao,
                   ROUND(TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao)/24.0,1)
            FROM chamados c LEFT JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 AND c.data_solucao IS NOT NULL
            ORDER BY TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao) DESC
        """)).fetchall()
        sheets.append({"title": "Tempo de Solução",
                        "headers": ["ID","Título","Empresa","Prioridade","Tipo","Abertura","Solução","Dias"],
                        "rows": [list(r) for r in rows7]})

        # Aba 8 - Ativos sem Contrato
        rows8 = (db.session.query(Ativo.nome, Empresa.nome, Localizacao.nome)
                 .join(Empresa, Ativo.empresa_id == Empresa.id)
                 .outerjoin(Localizacao, Ativo.localizacao_id == Localizacao.id)
                 .filter(Ativo.contrato_id == None)
                 .order_by(Empresa.nome, Ativo.nome).all())
        sheets.append({"title": "Ativos sem Contrato",
                        "headers": ["Ativo","Empresa","Localização"],
                        "rows": [[r[0], r[1], r[2] or 'Não informada'] for r in rows8]})

        buf = _make_excel(sheets)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"relatorio_completo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500
