# -*- coding: utf-8 -*-
from flask import Blueprint, request, jsonify, Response, send_file
from datetime import datetime, timedelta
import io
import re
import ssl
import urllib.request

from .. import db
from sqlalchemy import text
from ..models.contador_impressora import ContadorImpressora

# ============================================================
# Blueprint (TEM QUE EXISTIR ANTES DE QUALQUER @route)
# ============================================================
contador_impressora_bp = Blueprint('contador_impressora_bp', __name__)

# ============================================================
# CORS / OPTIONS
# ============================================================

@contador_impressora_bp.after_request
def after_request(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS, PATCH'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-API-Token'
    return response

@contador_impressora_bp.route('/<path:any>', methods=['OPTIONS'])
@contador_impressora_bp.route('/', methods=['OPTIONS'])
def handle_options(any=None):
    r = Response()
    r.headers['Access-Control-Allow-Origin'] = '*'
    r.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS, PATCH'
    r.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-API-Token'
    return r, 200


# ============================================================
# Helpers
# ============================================================

def _norm_lower(s):
    return (s or '').strip().lower()

def _to_int(x):
    try:
        if x is None:
            return None
        if isinstance(x, int):
            return x
        s = str(x).strip()
        s = s.replace('.', '').replace(',', '')
        return int(s) if s.lstrip('-').isdigit() else None
    except Exception:
        return None

def _extract_pct(s):
    if s is None:
        return None
    m = re.search(r'(\d{1,3})\s*%?', str(s))
    if not m:
        return None
    v = int(m.group(1))
    if 0 <= v <= 100:
        return v
    return None

def _http_get(url, verify_ssl=False, timeout=12, usuario=None, senha=None):
    """
    HTTP GET simples com User-Agent.
    Para HTTPS, por padrão NÃO valida certificado (muitos devices usam cert self-signed).
    Suporta Basic Auth via usuario/senha.
    """
    import base64 as _b64
    headers = {'User-Agent': 'Mozilla/5.0'}
    if usuario and senha:
        cred = _b64.b64encode(f'{usuario}:{senha}'.encode()).decode()
        headers['Authorization'] = f'Basic {cred}'
    req = urllib.request.Request(url, headers=headers)

    if url.lower().startswith('https'):
        if verify_ssl:
            return urllib.request.urlopen(req, timeout=timeout).read()
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        return urllib.request.urlopen(req, context=ctx, timeout=timeout).read()

    return urllib.request.urlopen(req, timeout=timeout).read()


# ============================================================
# Modelo (robusto + auto-detecção)
# ============================================================

def _get_modelo_tipo(c: ContadorImpressora):
    """
    Determina o modelo_tipo de forma robusta.
    1) usa campo modelo_tipo (se existir e estiver preenchido)
    2) tenta inferir por modelo + nome (muitos cadastros vêm com "Xerox7855" no nome)
    """
    modelo_tipo = getattr(c, 'modelo_tipo', None)
    if modelo_tipo:
        return _norm_lower(modelo_tipo)

    modelo = _norm_lower(getattr(c, 'modelo', None))
    nome = _norm_lower(getattr(c, 'nome', None))
    txt = f"{modelo} {nome}"

    if 'altalink' in txt:
        return 'xerox_altalink'
    if 'primelink' in txt:
        return 'xerox_primelink'

    # WorkCentre 7855/7835 (cadastros variam MUITO)
    # exemplos: "Xerox7855", "WC7855", "WorkCentre 7855", "7835"
    if (
        'workcentre' in txt
        or 'work centre' in txt
        or 'xerox7855' in txt
        or '7855' in txt
        or '7835' in txt
    ):
        return 'xerox_workcentre_7855_7835'

    # HP X557xx
    if 'x557' in txt:
        return 'hp_x557'

    return 'desconhecido'


# ============================================================
# Parsers
# ============================================================

def parse_xerox_altalink(ip: str):
    base = f'https://{ip}'
    url_insumos = f'{base}/stat/consumables.php'
    url_contadores = f'{base}/counters/usage.php'

    result = {'contadores': {}, 'insumos': {}}

    # INSUMOS
    html = _http_get(url_insumos, verify_ssl=False).decode('utf-8', errors='ignore')
    html_limpo = re.sub(r'<script.*?>.*?</script>', '', html, flags=re.DOTALL)
    linhas = re.findall(r'<tr class="(?:odd|even)">(.*?)</tr>', html_limpo, re.DOTALL)

    insumos = {}
    for linha in linhas:
        tds = re.findall(r'<td[^>]*>(.*?)</td>', linha, re.DOTALL)
        tds_txt = [re.sub(r'<[^>]+>', '', td).replace('&nbsp;', '').strip() for td in tds]
        tds_txt = [t for t in tds_txt if t]
        if len(tds_txt) >= 2:
            nome = tds_txt[0]
            nivel = next((t for t in tds_txt if '%' in t), None)
            pct = _extract_pct(nivel)
            if pct is not None:
                insumos[nome] = pct
    result['insumos'] = insumos

    # CONTADORES
    html_cont = _http_get(url_contadores, verify_ssl=False).decode('utf-8', errors='ignore')
    filtro = set([
        'Total Impressions', 'Color Impressions', 'Black Impressions',
        'A4 Equivalent Impressions', 'Large Impressions',
        'Color Copied Impressions', 'Color Printed Impressions',
        'Black Copied Impressions', 'Black Printed Impressions',
        'Black Large Impressions', 'Color Large Impressions'
    ])

    cont = {}
    linhas = re.findall(r'<tr[^>]*>(.*?)</tr>', html_cont, re.DOTALL)
    for linha in linhas:
        tds = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', linha, re.DOTALL)
        tds_txt = [re.sub(r'<[^>]+>', '', td).replace('&nbsp;', '').strip() for td in tds]
        tds_txt = [t for t in tds_txt if t]
        if len(tds_txt) >= 2 and tds_txt[0] in filtro:
            cont[tds_txt[0]] = _to_int(tds_txt[-1])
    result['contadores'] = cont

    return result


def parse_xerox_primelink(ip: str, http_usuario=None, http_senha=None, http_porta=None):
    """
    Suporta:
    - Porta 8080 HTTP (PrimeLink C9070 padrão)
    - Porta 80 HTTP (C60/C70 sem porta 8080, ex: São Lucas)
    """
    result = {'contadores': {}, 'insumos': {}}
    filtro = set([
        'Total Impressions','Black Impressions','Color Impressions',
        'Large Impressions','Black Large Impressions','Color Large Impressions',
        'Black Copied Impressions','Black Printed Impressions',
        'Color Copied Impressions','Color Printed Impressions'
    ])

    # Tentar bases em ordem
    if http_porta:
        proto = 'https' if int(http_porta) in (443, 8443) else 'http'
        bases = [f'{proto}://{ip}:{http_porta}']
    else:
        bases = [f'http://{ip}:8080', f'https://{ip}:8443', f'http://{ip}']
    html_prcnt = None
    html_stsply = None
    last_err = None

    for base in bases:
        try:
            h = _http_get(f'{base}/prcnt.htm', verify_ssl=False, timeout=8, usuario=http_usuario, senha=http_senha).decode('utf-8', errors='ignore')
            # Aceita se tiver var info= ou a lista de contadores
            if h and ('var info=' in h or 'Total Impressions' in h):
                html_prcnt = h
                try:
                    html_stsply = _http_get(f'{base}/stsply.htm', verify_ssl=False, timeout=8, usuario=http_usuario, senha=http_senha).decode('utf-8', errors='ignore')
                except Exception:
                    pass
                break
        except Exception as e:
            last_err = e
            continue

    if not html_prcnt:
        raise RuntimeError(f'PrimeLink {ip}: não foi possível acessar prcnt.htm (portas 8080/80). Último erro: {last_err}')

    # Parse contadores — dois formatos suportados:
    # Formato A (C9070 8080): var info=[('name',val), ...]
    # Formato B (C60 porta 80): var info=['name',val,'name2',val2,...]
    cont = {}
    m = re.search(r"var info=\[([^;]{10,}?)\]", html_prcnt, re.DOTALL)
    if m:
        items = re.findall(r"'([^']+)',\s*(\d+)", m.group(1))
        for nome, val in items:
            if nome in filtro:
                cont[nome] = _to_int(val)
    result['contadores'] = cont

    # Parse suprimentos
    insumos = {}
    if html_stsply:
        suprimentos = re.findall(r"\['([^']*)',\s*\d+,\s*(\d+)\]", html_stsply)
        # Para K1/K2 (dois cartuchos preto), guardar o maior valor
        _tmp = {}
        for item, level in suprimentos:
            if not item:
                continue
            pct = _extract_pct(level)
            if pct is None:
                continue
            nome = item.strip()
            # Se já existe variante do mesmo suprimento (ex: K1 vs K2), guarda o maior
            base = re.sub(r'\s*\[K\d\]', '', nome).strip()
            if base in _tmp:
                _tmp[base] = max(_tmp[base], pct)
            else:
                _tmp[base] = pct
            insumos[nome] = pct  # manter nome original também
        # Sobrescrever com o maior entre variantes
        for base, val in _tmp.items():
            insumos[base] = val
    result['insumos'] = insumos

    return result


def parse_xerox_workcentre_7855_7835(ip: str):
    """
    WorkCentre 7855/7835:
      - Contadores: /counters/billing_info.php (http) (fallback /counters/usage.php)
      - Suprimentos/Drums: /stat/consumables.php (muito comum)
    """
    base_http = f'http://{ip}'
    base_https = f'https://{ip}'

    urls_contadores = [
        f'{base_http}/counters/billing_info.php',
        f'{base_https}/counters/billing_info.php',
        f'{base_http}/counters/usage.php',
        f'{base_https}/counters/usage.php',
    ]

    urls_consumiveis = [
        f'{base_http}/stat/consumables.php',
        f'{base_https}/stat/consumables.php',
        # fallbacks (alguns firmwares têm)
        f'{base_http}/status/consumables.php',
        f'{base_https}/status/consumables.php',
    ]

    result = {'contadores': {}, 'insumos': {}}

    # ---- contadores (obrigatório) ----
    html_cont = None
    last_err = None
    for url in urls_contadores:
        try:
            html_cont = _http_get(url, verify_ssl=False, timeout=8).decode('utf-8', errors='ignore')
            if html_cont and len(html_cont) > 50:
                break
        except Exception as e:
            last_err = e
            continue
    if not html_cont:
        raise RuntimeError(f'Não foi possível acessar contadores (último erro: {last_err})')

    numeros = re.findall(r'<td class="rightAlign[^>]*">([0-9]+)</td>', html_cont)
    if not numeros:
        numeros = re.findall(r'<td[^>]*>\s*([0-9]+)\s*</td>', html_cont)

    labels = [
        'Black Impressions',
        'Color Impressions',
        'Total Impressions',
        'Black Large Impressions',
        'Color Large Impressions'
    ]

    cont = {}
    for i, num in enumerate(numeros):
        if i < len(labels):
            cont[labels[i]] = _to_int(num)
    result['contadores'] = cont

    # ---- consumíveis (opcional) ----
    html_cons = None
    for url in urls_consumiveis:
        try:
            html_cons = _http_get(url, verify_ssl=False, timeout=8).decode('utf-8', errors='ignore')
            if html_cons and len(html_cons) > 50:
                break
        except Exception:
            continue

    insumos = {}
    if html_cons:
        html_cons_limpo = re.sub(r'<script.*?>.*?</script>', '', html_cons, flags=re.DOTALL)

        linhas = re.findall(r'<tr class="(?:odd|even)">(.*?)</tr>', html_cons_limpo, re.DOTALL)
        if not linhas:
            linhas = re.findall(r'<tr[^>]*>(.*?)</tr>', html_cons_limpo, re.DOTALL)

        for linha in linhas:
            tds = re.findall(r'<td[^>]*>(.*?)</td>', linha, re.DOTALL)
            tds_txt = [re.sub(r'<[^>]+>', '', td).replace('&nbsp;', '').strip() for td in tds]
            tds_txt = [t for t in tds_txt if t]
            if len(tds_txt) >= 2:
                nome = tds_txt[0]
                nivel = next((t for t in tds_txt if '%' in t), None)
                pct = _extract_pct(nivel)
                if pct is not None:
                    insumos[nome] = pct

        # fallback genérico
        if not insumos:
            pares = re.findall(r'>([^<>]{2,60})<.*?([0-9]{1,3})\s*%', html_cons_limpo, re.DOTALL)
            for nome, level in pares:
                nome = re.sub(r'\s+', ' ', nome).strip()
                pct = _extract_pct(level)
                if nome and pct is not None:
                    insumos[nome] = pct

    result['insumos'] = insumos
    return result


def parse_hp_x557(ip: str):
    base = f'https://{ip}'
    url_principal = f'{base}'
    url_contador  = f'{base}/hp/device/InternalPages/Index?id=UsagePage'

    result = {'contadores': {}, 'insumos': {}}

    html_cont = _http_get(url_contador, verify_ssl=False).decode('utf-8', errors='ignore')
    alvos = {
        'A4 Monochrome': r'Print\.A4\.Monochrome\" class=\"align-right\">([0-9,]+)',
        'A4 Color':      r'Print\.A4\.Color\" class=\"align-right\">([0-9,]+)',
        'A4 Total':      r'Print\.A4\.Total\" class=\"align-right\">([0-9,]+)'
    }

    cont = {}
    for nome, regex in alvos.items():
        m = re.search(regex, html_cont)
        if m:
            cont[nome] = _to_int(m.group(1))
    result['contadores'] = cont

    html_home = _http_get(url_principal, verify_ssl=False).decode('utf-8', errors='ignore')
    insumos = {}
    for i in range(5):
        regex_nome = rf'SupplyName{i}\" title=\"([^\"]+)'
        regex_level = rf'SupplyGauge{i}\" style=\"width:([0-9]+)'
        m_nome = re.search(regex_nome, html_home)
        m_level = re.search(regex_level, html_home)
        if m_nome and m_level:
            nome_insumo = m_nome.group(1).strip()
            pct = _extract_pct(m_level.group(1))
            if pct is not None:
                insumos[nome_insumo] = pct
    result['insumos'] = insumos

    return result


def coletar_por_modelo(modelo_tipo: str, ip: str, http_usuario=None, http_senha=None, http_porta=None):
    mt = _norm_lower(modelo_tipo)

    if mt == 'xerox_altalink':
        return parse_xerox_altalink(ip)
    if mt == 'xerox_primelink':
        return parse_xerox_primelink(ip, http_usuario=http_usuario, http_senha=http_senha, http_porta=http_porta)
    if mt == 'xerox_workcentre_7855_7835':
        return parse_xerox_workcentre_7855_7835(ip)
    if mt == 'hp_x557':
        return parse_hp_x557(ip)

    # AUTO-DETECT (quando cadastro vier "desconhecido")
    if mt in ('', 'desconhecido', 'unknown', 'none', 'null'):
        # 1) WorkCentre 7855/7835 -> billing_info.php
        try:
            _http_get(f'http://{ip}/counters/billing_info.php', verify_ssl=False, timeout=5)
            return parse_xerox_workcentre_7855_7835(ip)
        except Exception:
            pass

        # 2) AltaLink -> counters/usage.php em https
        try:
            _http_get(f'https://{ip}/counters/usage.php', verify_ssl=False, timeout=5)
            return parse_xerox_altalink(ip)
        except Exception:
            pass

        # 3) PrimeLink -> porta 8080 /prcnt.htm
        try:
            _http_get(f'http://{ip}:8080/prcnt.htm', verify_ssl=True, timeout=5)
            return parse_xerox_primelink(ip, http_usuario=http_usuario, http_senha=http_senha, http_porta=http_porta)
        except Exception:
            pass

    raise ValueError(f"modelo_tipo não suportado: {modelo_tipo}")


# ============================================================
# Mapeamento para o model
# ============================================================

def _map_contadores_e_insumos_para_model(c: ContadorImpressora, payload: dict):
    cont = payload.get('contadores') or {}
    ins = payload.get('insumos') or {}

    # Xerox (gerais)
    if cont.get('Total Impressions') is not None:
        c.contador_total = _to_int(cont.get('Total Impressions'))
    if cont.get('Black Impressions') is not None:
        c.contador_pb = _to_int(cont.get('Black Impressions'))
    if cont.get('Color Impressions') is not None:
        c.contador_color = _to_int(cont.get('Color Impressions'))

    # A3 Xerox
    if cont.get('Black Large Impressions') is not None:
        c.contador_a3_pb = _to_int(cont.get('Black Large Impressions'))
    if cont.get('Color Large Impressions') is not None:
        c.contador_a3_color = _to_int(cont.get('Color Large Impressions'))

    # HP A4
    if cont.get('A4 Total') is not None:
        c.contador_total = _to_int(cont.get('A4 Total'))
    if cont.get('A4 Monochrome') is not None:
        c.contador_a4_pb = _to_int(cont.get('A4 Monochrome'))
    if cont.get('A4 Color') is not None:
        c.contador_a4_color = _to_int(cont.get('A4 Color'))

    # insumos - heurística por palavra
    def set_if_match(attr, *keys):
        vals = []
        for nome, pct in ins.items():
            n = _norm_lower(nome)
            if any(k in n for k in keys):
                p = _extract_pct(pct)
                if p is not None:
                    vals.append(p)
        if vals:
            setattr(c, attr, min(vals))

    set_if_match('toner_preto_nivel', 'preto', 'black', 'bk', 'toner black', 'k')
    set_if_match('toner_ciano_nivel', 'ciano', 'cyan')
    set_if_match('toner_magenta_nivel', 'magenta')
    set_if_match('toner_amarelo_nivel', 'amarelo', 'yellow')
    set_if_match('reservatorio_nivel', 'waste', 'residu', 'reserv', 'container')

    # drums (heurística; sem mapear R1..R4 por cor)
    if hasattr(c, 'drum_preto_nivel'):
        set_if_match('drum_preto_nivel', 'drum', 'black drum', 'drum black', 'k drum')
    if hasattr(c, 'drum_ciano_nivel'):
        set_if_match('drum_ciano_nivel', 'drum', 'cyan drum', 'drum cyan')
    if hasattr(c, 'drum_magenta_nivel'):
        set_if_match('drum_magenta_nivel', 'drum', 'magenta drum', 'drum magenta')
    if hasattr(c, 'drum_amarelo_nivel'):
        set_if_match('drum_amarelo_nivel', 'drum', 'yellow drum', 'drum yellow')

    if hasattr(c, 'suprimentos_raw'):
        raw_list = []
        for nome, pct in ins.items():
            raw_list.append({
                'categoria': 'web',
                'descricao': nome,
                'percentual': _extract_pct(pct)
            })
        c.suprimentos_raw = raw_list

    c.status = 'online'
    c.ultima_leitura = datetime.utcnow()


# ============================================================
# ROTAS CRUD
# ============================================================

@contador_impressora_bp.route('/', methods=['GET'])
def listar():
    empresa_id = request.args.get('empresa_id')
    q = ContadorImpressora.query
    if empresa_id:
        q = q.filter_by(empresa_id=int(empresa_id))
    return jsonify([c.to_dict() for c in q.all()])


@contador_impressora_bp.route('/<int:id>', methods=['GET'])
def get_one(id):
    c = ContadorImpressora.query.get_or_404(id)
    return jsonify(c.to_dict())


@contador_impressora_bp.route('/', methods=['POST'])
def criar():
    data = request.get_json() or {}
    novo = ContadorImpressora(
        nome=data['nome'],
        ip=data['ip'],
        community=data.get('community', 'public'),
        modelo=data.get('modelo'),
        numero_serie=data.get('numero_serie'),
        localizacao=data.get('localizacao'),
        empresa_id=data.get('empresa_id'),
    )
    if hasattr(novo, 'modelo_tipo'):
        setattr(novo, 'modelo_tipo', data.get('modelo_tipo'))

    db.session.add(novo)
    db.session.commit()
    return jsonify(novo.to_dict()), 201


@contador_impressora_bp.route('/<int:id>', methods=['PUT'])
def atualizar(id):
    c = ContadorImpressora.query.get_or_404(id)
    data = request.get_json() or {}

    c.nome = data.get('nome', c.nome)
    c.ip = data.get('ip', c.ip)
    c.community = data.get('community', c.community)
    c.modelo = data.get('modelo', c.modelo)
    c.numero_serie = data.get('numero_serie', c.numero_serie)
    c.localizacao = data.get('localizacao', c.localizacao)
    c.empresa_id = data.get('empresa_id', c.empresa_id)

    if hasattr(c, 'modelo_tipo') and 'modelo_tipo' in data:
        setattr(c, 'modelo_tipo', data.get('modelo_tipo'))

    db.session.commit()
    return jsonify(c.to_dict())


@contador_impressora_bp.route('/<int:id>', methods=['DELETE'])
def deletar(id):
    c = ContadorImpressora.query.get_or_404(id)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'success': True})


# ============================================================
# CONSULTA (mantém /consultar-snmp por compatibilidade)
# ============================================================

@contador_impressora_bp.route('/<int:id>/consultar-snmp', methods=['POST'])
def consultar(id):
    c = ContadorImpressora.query.get_or_404(id)
    try:
        modelo_tipo = _get_modelo_tipo(c)
        payload = coletar_por_modelo(
            modelo_tipo, c.ip,
            http_usuario=getattr(c, 'http_usuario', None),
            http_senha=getattr(c, 'http_senha', None),
            http_porta=getattr(c, 'http_porta', None)
        )
        _map_contadores_e_insumos_para_model(c, payload)

        # Salvar snapshot no histórico
        import json as _json
        db.session.execute(text("""
            INSERT INTO contadores_impressora_historico
              (impressora_id, lido_em,
               contador_total, contador_pb, contador_color,
               contador_a3_pb, contador_a3_color,
               contador_a4_pb, contador_a4_color,
               toner_preto_nivel, toner_ciano_nivel,
               toner_magenta_nivel, toner_amarelo_nivel,
               reservatorio_nivel,
               drum_preto_nivel, drum_ciano_nivel,
               drum_magenta_nivel, drum_amarelo_nivel,
               suprimentos_raw)
            VALUES
              (:imp_id, :lido_em,
               :total, :pb, :color,
               :a3pb, :a3color,
               :a4pb, :a4color,
               :tpk, :tc, :tm, :ty,
               :res,
               :dpk, :dc, :dm, :dy,
               :raw)
        """), {
            'imp_id':  c.id,
            'lido_em': datetime.utcnow(),
            'total':   c.contador_total,
            'pb':      c.contador_pb,
            'color':   c.contador_color,
            'a3pb':    c.contador_a3_pb,
            'a3color': c.contador_a3_color,
            'a4pb':    getattr(c, 'contador_a4_pb', None),
            'a4color': getattr(c, 'contador_a4_color', None),
            'tpk':     c.toner_preto_nivel,
            'tc':      c.toner_ciano_nivel,
            'tm':      c.toner_magenta_nivel,
            'ty':      c.toner_amarelo_nivel,
            'res':     c.reservatorio_nivel,
            'dpk':     c.drum_preto_nivel,
            'dc':      c.drum_ciano_nivel,
            'dm':      c.drum_magenta_nivel,
            'dy':      c.drum_amarelo_nivel,
            'raw':     _json.dumps(c.suprimentos_raw or []),
        })

        db.session.commit()
        return jsonify({'success': True, 'status': 'online', 'dados': c.to_dict()})
    except Exception as e:
        c.status = 'offline'
        c.ultima_leitura = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': False, 'status': 'offline', 'mensagem': str(e)}), 200


@contador_impressora_bp.route('/atualizar-todas', methods=['POST'])
def atualizar_todas():
    import os
    import secrets

    token_recebido = request.headers.get(
        'X-Cron-Token',
        ''
    )
    token_configurado = os.environ.get(
        'CRON_TOKEN',
        ''
    )

    if (
        not token_configurado
        or not token_recebido
        or not secrets.compare_digest(
            token_recebido,
            token_configurado
        )
    ):
        return jsonify({
            'error': 'Token de cron inválido'
        }), 403

    resultados = {
        'sucesso': 0,
        'falha': 0,
        'detalhes': []
    }

    impressoras = ContadorImpressora.query.all()

    for c in impressoras:
        try:
            modelo_tipo = _get_modelo_tipo(c)
            payload = coletar_por_modelo(
            modelo_tipo, c.ip,
            http_usuario=getattr(c, 'http_usuario', None),
            http_senha=getattr(c, 'http_senha', None),
            http_porta=getattr(c, 'http_porta', None)
        )
            _map_contadores_e_insumos_para_model(c, payload)

            # Salvar snapshot no histórico
            import json as _json
            db.session.execute(text("""
                INSERT INTO contadores_impressora_historico
                  (impressora_id, lido_em,
                   contador_total, contador_pb, contador_color,
                   contador_a3_pb, contador_a3_color,
                   contador_a4_pb, contador_a4_color,
                   toner_preto_nivel, toner_ciano_nivel,
                   toner_magenta_nivel, toner_amarelo_nivel,
                   reservatorio_nivel,
                   drum_preto_nivel, drum_ciano_nivel,
                   drum_magenta_nivel, drum_amarelo_nivel,
                   suprimentos_raw)
                VALUES
                  (:imp_id, :lido_em,
                   :total, :pb, :color,
                   :a3pb, :a3color,
                   :a4pb, :a4color,
                   :tpk, :tc, :tm, :ty,
                   :res,
                   :dpk, :dc, :dm, :dy,
                   :raw)
            """), {
                'imp_id':  c.id,
                'lido_em': datetime.utcnow(),
                'total':   c.contador_total,
                'pb':      c.contador_pb,
                'color':   c.contador_color,
                'a3pb':    c.contador_a3_pb,
                'a3color': c.contador_a3_color,
                'a4pb':    getattr(c, 'contador_a4_pb', None),
                'a4color': getattr(c, 'contador_a4_color', None),
                'tpk':     c.toner_preto_nivel,
                'tc':      c.toner_ciano_nivel,
                'tm':      c.toner_magenta_nivel,
                'ty':      c.toner_amarelo_nivel,
                'res':     c.reservatorio_nivel,
                'dpk':     c.drum_preto_nivel,
                'dc':      c.drum_ciano_nivel,
                'dm':      c.drum_magenta_nivel,
                'dy':      c.drum_amarelo_nivel,
                'raw':     _json.dumps(getattr(c, 'suprimentos_raw', None) or []),
            })
            resultados['sucesso'] += 1
            resultados['detalhes'].append({'id': c.id, 'nome': c.nome, 'status': 'online'})
        except Exception as e:
            c.status = 'offline'
            c.ultima_leitura = datetime.utcnow()
            resultados['falha'] += 1
            resultados['detalhes'].append({'id': c.id, 'nome': c.nome, 'status': 'offline', 'erro': str(e)})

    db.session.commit()
    return jsonify(resultados)



# ============================================================
# EXPORTAR EXCEL
# ============================================================


@contador_impressora_bp.route('/exportar-excel', methods=['GET'])
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from sqlalchemy import text as _text
    from collections import OrderedDict
    from datetime import date as _date, timedelta as _td

    empresa_id  = request.args.get('empresa_id')
    data_inicio = request.args.get('data_inicio')  # YYYY-MM-DD
    data_fim    = request.args.get('data_fim')      # YYYY-MM-DD

    query = ContadorImpressora.query
    if empresa_id:
        query = query.filter_by(empresa_id=int(empresa_id))
    impressoras = query.order_by(ContadorImpressora.nome).all()
    imp_ids = [i.id for i in impressoras]

    # --- filtro de datas ---
    where_data = ""
    params = {}
    if data_inicio:
        where_data += " AND DATE(lido_em) >= :di"
        params['di'] = data_inicio
    if data_fim:
        where_data += " AND DATE(lido_em) <= :df"
        params['df'] = data_fim

    # --- buscar última leitura de cada impressora por dia ---
    historico_rows = []
    if imp_ids:
        ids_str = ','.join(str(x) for x in imp_ids)
        sql = f"""
            SELECT h.impressora_id,
                   DATE(h.lido_em) as data_leitura,
                   h.contador_total, h.contador_pb, h.contador_color,
                   h.contador_a3_pb, h.contador_a3_color,
                   h.contador_a4_pb, h.contador_a4_color,
                   h.toner_preto_nivel, h.toner_ciano_nivel,
                   h.toner_magenta_nivel, h.toner_amarelo_nivel,
                   h.reservatorio_nivel,
                   h.drum_preto_nivel, h.drum_ciano_nivel,
                   h.drum_magenta_nivel, h.drum_amarelo_nivel
            FROM contadores_impressora_historico h
            INNER JOIN (
                SELECT impressora_id, DATE(lido_em) as d, MAX(lido_em) as max_lido
                FROM contadores_impressora_historico
                WHERE impressora_id IN ({ids_str}) {where_data}
                GROUP BY impressora_id, DATE(lido_em)
            ) sub ON h.impressora_id = sub.impressora_id AND h.lido_em = sub.max_lido
            WHERE h.impressora_id IN ({ids_str})
            ORDER BY data_leitura ASC, h.impressora_id ASC
        """
        historico_rows = db.session.execute(_text(sql), params).fetchall()

    # --- montar pivot[data_str][imp_id] = {...} ---
    pivot = OrderedDict()
    for row in historico_rows:
        imp_id   = row[0]
        data_obj = row[1]
        if hasattr(data_obj, 'strftime'):
            data_str = data_obj.strftime('%d/%m/%Y')
        else:
            y, m, d = str(data_obj).split('-')
            data_str = f"{d}/{m}/{y}"
        if data_str not in pivot:
            pivot[data_str] = {}
        pivot[data_str][imp_id] = {
            'total':         row[2],
            'pb':            row[3],
            'color':         row[4],
            'a3pb':          row[5],
            'a3color':       row[6],
            'a4pb':          row[7],
            'a4color':       row[8],
            'toner_preto':   row[9],
            'toner_ciano':   row[10],
            'toner_magenta': row[11],
            'toner_amarelo': row[12],
            'reservatorio':  row[13],
            'drum_preto':    row[14],
            'drum_ciano':    row[15],
            'drum_magenta':  row[16],
            'drum_amarelo':  row[17],
        }

    # --- gerar range completo de datas do período ---
    def parse_br(s):
        d, m, y = s.split('/')
        return _date(int(y), int(m), int(d))

    if pivot:
        data_min = min(parse_br(d) for d in pivot.keys())
        data_max = max(parse_br(d) for d in pivot.keys())
        if data_inicio:
            try:
                y, m, d = data_inicio.split('-')
                data_min = max(data_min, _date(int(y), int(m), int(d)))
            except: pass
        if data_fim:
            try:
                y, m, d = data_fim.split('-')
                data_max = min(data_max, _date(int(y), int(m), int(d)))
            except: pass
        todas_datas = []
        cur = data_min
        while cur <= data_max:
            todas_datas.append(cur.strftime('%d/%m/%Y'))
            cur += _td(days=1)
    else:
        todas_datas = []

    campos = [
        ('Total',           'total'),
        ('PB',              'pb'),
        ('Color',           'color'),
        ('A3 PB',           'a3pb'),
        ('A3 Color',        'a3color'),
        ('A4 PB',           'a4pb'),
        ('A4 Color',        'a4color'),
        ('Toner Preto %',   'toner_preto'),
        ('Toner Ciano %',   'toner_ciano'),
        ('Toner Magenta %', 'toner_magenta'),
        ('Toner Amarelo %', 'toner_amarelo'),
        ('Reserv. %',       'reservatorio'),
        ('Drum Preto %',    'drum_preto'),
        ('Drum Ciano %',    'drum_ciano'),
        ('Drum Magenta %',  'drum_magenta'),
        ('Drum Amarelo %',  'drum_amarelo'),
    ]
    n_campos   = len(campos)
    n_imp      = len(impressoras)
    total_cols = 1 + n_imp * n_campos

    imp_fills_header = [
        PatternFill('solid', fgColor='1F2937'),
        PatternFill('solid', fgColor='1E3A5F'),
        PatternFill('solid', fgColor='1A4A2E'),
        PatternFill('solid', fgColor='4A1E3A'),
        PatternFill('solid', fgColor='4A3A1E'),
        PatternFill('solid', fgColor='2E1E4A'),
        PatternFill('solid', fgColor='3A2E1E'),
        PatternFill('solid', fgColor='1E3A3A'),
    ]
    imp_fills_data = [
        PatternFill('solid', fgColor='EFF6FF'),
        PatternFill('solid', fgColor='F0FFF4'),
        PatternFill('solid', fgColor='FFF0F6'),
        PatternFill('solid', fgColor='FFFFF0'),
        PatternFill('solid', fgColor='F5F0FF'),
        PatternFill('solid', fgColor='F0FFFF'),
        PatternFill('solid', fgColor='FFF5F0'),
        PatternFill('solid', fgColor='F0F5FF'),
    ]
    offline_fill = PatternFill('solid', fgColor='F3F4F6')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contadores por Data'

    periodo = ''
    if data_inicio and data_fim:
        periodo = f' | Período: {data_inicio} a {data_fim}'
    elif data_inicio:
        periodo = f' | A partir de: {data_inicio}'
    elif data_fim:
        periodo = f' | Até: {data_fim}'

    titulo = f'Relatório de Contadores por Data{periodo} — gerado em {(datetime.utcnow() - timedelta(hours=3)).strftime("%d/%m/%Y %H:%M")}'
    ws.cell(row=1, column=1, value=titulo)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Linha 2 — Data (merge com linha 3)
    ws.cell(row=2, column=1, value='Data')
    ws.cell(row=2, column=1).fill = PatternFill('solid', fgColor='1F2937')
    ws.cell(row=2, column=1).font = Font(bold=True, color='FFFFFF')
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    # Linha 2 — nome das impressoras
    for idx, imp in enumerate(impressoras):
        col_start = 2 + idx * n_campos
        col_end   = col_start + n_campos - 1
        fill      = imp_fills_header[idx % len(imp_fills_header)]
        ws.cell(row=2, column=col_start, value=imp.nome)
        ws.cell(row=2, column=col_start).fill = fill
        ws.cell(row=2, column=col_start).font = Font(bold=True, color='FFFFFF', size=10)
        ws.cell(row=2, column=col_start).alignment = Alignment(horizontal='center', vertical='center')
        if col_end > col_start:
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        # Linha 3 — sub-cabeçalhos
        for j, (nome_campo, _) in enumerate(campos):
            c = ws.cell(row=3, column=col_start + j, value=nome_campo)
            c.fill = fill
            c.font = Font(bold=True, color='FFFFFF', size=8)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 32
    ws.column_dimensions['A'].width = 12
    for idx in range(n_imp):
        col_start = 2 + idx * n_campos
        for j in range(n_campos):
            ws.column_dimensions[get_column_letter(col_start + j)].width = 11

    # --- Dados: uma linha por data, todas as impressoras ---
    row_num = 4
    for data_str in todas_datas:
        dados_dia = pivot.get(data_str, {})
        ws.cell(row=row_num, column=1, value=data_str)
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

        for idx, imp in enumerate(impressoras):
            col_start = 2 + idx * n_campos
            if imp.id in dados_dia:
                fill_data = imp_fills_data[idx % len(imp_fills_data)]
                leitura   = dados_dia[imp.id]
                for j, (_, chave) in enumerate(campos):
                    cell = ws.cell(row=row_num, column=col_start + j, value=leitura.get(chave))
                    cell.fill = fill_data
                    cell.alignment = Alignment(horizontal='center')
            else:
                for j in range(n_campos):
                    cell = ws.cell(row=row_num, column=col_start + j, value='—')
                    cell.fill = offline_fill
                    cell.alignment = Alignment(horizontal='center')
        row_num += 1

    # --- Aba Resumo Atual ---
    ws2 = wb.create_sheet('Resumo Atual')
    resumo_headers = ['Nome', 'IP', 'Modelo', 'Status', 'Última Leitura',
                      'Total', 'PB', 'Color',
                      'Toner Preto %', 'Toner Ciano %', 'Toner Magenta %', 'Toner Amarelo %']
    for col, h in enumerate(resumo_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = PatternFill('solid', fgColor='1F2937')
        c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(col)].width = 18

    for r, imp in enumerate(impressoras, 2):
        ul = getattr(imp, 'ultima_leitura', None)
        ws2.cell(row=r, column=1,  value=imp.nome)
        ws2.cell(row=r, column=2,  value=imp.ip)
        ws2.cell(row=r, column=3,  value=getattr(imp, 'modelo', None))
        ws2.cell(row=r, column=4,  value=getattr(imp, 'status', None))
        ws2.cell(row=r, column=5,  value=(ul - timedelta(hours=3)).strftime('%d/%m/%Y %H:%M') if ul else None)
        ws2.cell(row=r, column=6,  value=getattr(imp, 'contador_total', None))
        ws2.cell(row=r, column=7,  value=getattr(imp, 'contador_pb', None))
        ws2.cell(row=r, column=8,  value=getattr(imp, 'contador_color', None))
        ws2.cell(row=r, column=9,  value=getattr(imp, 'toner_preto_nivel', None))
        ws2.cell(row=r, column=10, value=getattr(imp, 'toner_ciano_nivel', None))
        ws2.cell(row=r, column=11, value=getattr(imp, 'toner_magenta_nivel', None))
        ws2.cell(row=r, column=12, value=getattr(imp, 'toner_amarelo_nivel', None))
        if getattr(imp, 'status', None) == 'offline':
            for col in range(1, 13):
                ws2.cell(row=r, column=col).fill = PatternFill('solid', fgColor='FEE2E2')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name='contadores_impressora.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
