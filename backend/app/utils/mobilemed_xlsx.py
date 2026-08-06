# -*- coding: utf-8 -*-
"""Conversao e consolidacao dos relatorios MobileMed em XLSX."""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import unicodedata
import uuid
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BACKEND_DIR = Path(__file__).resolve().parents[2]
MOBILEMED_STORAGE_DIR = Path(
    os.environ.get(
        "MOBILEMED_STORAGE_DIR",
        str(BACKEND_DIR / "instance" / "mobilemed_reports"),
    )
)
MOBILEMED_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
BI_XLSX_PATH = MOBILEMED_STORAGE_DIR / "BI_MobileMed_Consolidado.xlsx"


FIELD_ALIASES = {
    "exame.id": ["exame.id", "exame_id", "id_exame", "id exame"],
    "exame.nome_paciente": [
        "exame.nome_paciente", "nome_paciente", "nome paciente",
        "nome do paciente", "paciente",
    ],
    "exame.data_criacao": [
        "exame.data_criacao", "data_criacao", "data criacao",
        "data de criacao",
    ],
    "exame.data_realizacao": [
        "exame.data_realizacao", "data_realizacao", "data realizacao",
        "data de realizacao",
    ],
    "exame.codigo_pedido": [
        "exame.codigo_pedido", "codigo_pedido", "codigo pedido", "pedido",
    ],
    "exame.pacs_accession_no": [
        "exame.pacs_accession_no", "pacs_accession_no", "accession",
        "accession number",
    ],
    "empresa.nome_fantasia": [
        "empresa.nome_fantasia",
        "empresa_nome_fantasia",
        "nome_fantasia_empresa",
        "nome fantasia da empresa",
        "nome da empresa",
        "empresa.nome",
        "empresa_nome",
        "nome_empresa",
        "nome_fantasia",
        "nome fantasia",
        "fantasia",
        "unidade.nome",
        "unidade_nome",
        "nome_unidade",
        "nome da unidade",
        "unidade",
        "unity.name",
        "unity_name",
        "company.name",
        "company_name",
        "empresa",
    ],
    "usuario.nome": [
        "usuario.nome", "usuario_nome", "nome usuario", "medico", "médico",
    ],
    "usuario.crm": ["usuario.crm", "usuario_crm", "crm"],
    "status.descricao": [
        "status.descricao", "status_descricao", "descricao_status", "status",
    ],
    "modalidade.nome": [
        "modalidade.nome", "modalidade_nome", "modalidade",
    ],
}


def is_bi_name(nome: str | None) -> bool:
    """Aceita BI, B.I, B-I e B_I no inicio do nome."""
    compacto = re.sub(r"[^A-Z0-9]", "", (nome or "").upper())
    return compacto.startswith("BI")


def normalize_selected_fields(
    campos: Any,
    nome_relatorio: str,
    campos_disponiveis: list[str],
) -> list[str]:
    """Valida, remove repeticoes e preserva a ordem escolhida."""
    if not isinstance(campos, list):
        campos = []

    resultado: list[str] = []
    for campo in campos:
        if not isinstance(campo, str):
            continue
        campo = campo.strip()
        if campo in campos_disponiveis and campo not in resultado:
            resultado.append(campo)

    if is_bi_name(nome_relatorio) and "exame.id" not in resultado:
        resultado.insert(0, "exame.id")

    if "usuario.nome" in resultado and "usuario.digitador_nome" in resultado:
        raise ValueError(
            "A MobileMed nao aceita usuario.nome junto com "
            "usuario.digitador_nome. Selecione apenas um deles."
        )

    if (
        "laudo_usuario.action" in resultado
        and any(campo.startswith("usuario.") for campo in resultado)
    ):
        raise ValueError(
            "A MobileMed pode falhar quando laudo_usuario.action e combinado "
            "com campos usuario.*."
        )

    return resultado


def safe_filename(nome: str | None) -> str:
    texto = unicodedata.normalize("NFKD", nome or "")
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^A-Za-z0-9._-]+", "_", texto).strip("._")
    return texto or "Relatorio_MobileMed"


def _report_csv_path(relatorio_id: int) -> Path:
    return MOBILEMED_STORAGE_DIR / f"mobilemed_{relatorio_id}.csv"


def _looks_like_csv_text(texto: str) -> bool:
    inicio = (texto or "").lstrip()[:1]
    return bool(inicio) and inicio not in ("{", "[")


def _atomic_write(path: Path, content: bytes) -> None:
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    temp.write_bytes(content)
    os.replace(temp, path)


def cache_report_csv(relatorio, force: bool = False) -> Path:
    """Baixa e guarda uma copia local para nao depender de URL temporaria."""
    destino = _report_csv_path(relatorio.id)
    if not force and destino.exists() and destino.stat().st_size > 0:
        return destino

    conteudo: bytes | None = None

    if relatorio.csv_dados and _looks_like_csv_text(relatorio.csv_dados):
        conteudo = relatorio.csv_dados.encode("utf-8")
    elif relatorio.csv_url:
        resposta = requests.get(
            relatorio.csv_url,
            timeout=(15, 300),
            allow_redirects=True,
        )
        resposta.raise_for_status()
        conteudo = resposta.content

    if not conteudo:
        raise ValueError(
            "O relatorio nao possui CSV salvo nem uma URL valida para download."
        )

    if conteudo[:2] == b"PK":
        with zipfile.ZipFile(io.BytesIO(conteudo)) as arquivo_zip:
            nomes_csv = [
                nome for nome in arquivo_zip.namelist()
                if nome.lower().endswith(".csv")
            ]
            if not nomes_csv:
                raise ValueError("O arquivo ZIP recebido nao contem um CSV.")
            conteudo = arquivo_zip.read(nomes_csv[0])

    _atomic_write(destino, conteudo)
    return destino


def _decode_csv(conteudo: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return conteudo.decode(encoding)
        except UnicodeDecodeError:
            continue
    return conteudo.decode("utf-8", errors="replace")


def _parse_csv(conteudo: bytes) -> tuple[list[str], list[dict[str, str]]]:
    texto = _decode_csv(conteudo)
    amostra = texto[:16384]

    try:
        dialect = csv.Sniffer().sniff(amostra, delimiters=",;|\t")
        delimitador = dialect.delimiter
    except csv.Error:
        delimitador = ";" if amostra.count(";") >= amostra.count(",") else ","

    leitor = csv.DictReader(io.StringIO(texto), delimiter=delimitador)
    cabecalhos = [str(c or "").strip() for c in (leitor.fieldnames or [])]
    if not cabecalhos:
        raise ValueError("Nao foi possivel identificar o cabecalho do CSV.")

    linhas: list[dict[str, str]] = []
    for linha in leitor:
        limpa = {
            str(chave).strip(): valor if valor is not None else ""
            for chave, valor in linha.items()
            if chave is not None
        }
        if any(str(valor).strip() for valor in limpa.values()):
            linhas.append(limpa)

    return cabecalhos, linhas


def _normalize_column(value: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(value or ""))
    texto = texto.encode("ascii", "ignore").decode("ascii")
    # Separa camelCase/PascalCase antes de normalizar.
    texto = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", texto)
    texto = texto.lower()
    return re.sub(r"[^a-z0-9]", "", texto)


def _column_tokens(value: Any) -> set[str]:
    texto = unicodedata.normalize("NFKD", str(value or ""))
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", texto)
    return {
        token
        for token in re.findall(r"[a-z0-9]+", texto.lower())
        if token not in {"da", "de", "do", "das", "dos", "the"}
    }


def _non_empty_count(cabecalho: str, linhas_csv: list[dict[str, str]] | None) -> int:
    if not linhas_csv:
        return 0
    return sum(
        1
        for linha in linhas_csv
        if str(linha.get(cabecalho, "") or "").strip()
    )


def _find_actual_header(
    campo: str,
    cabecalhos: list[str],
    linhas_csv: list[dict[str, str]] | None = None,
) -> str | None:
    """Localiza o cabecalho real, inclusive quando a API muda nome ou ordem."""
    candidatos = [
        campo,
        campo.replace(".", "_"),
        campo.split(".")[-1],
        *FIELD_ALIASES.get(campo, []),
    ]
    candidatos_normalizados = {
        _normalize_column(candidato)
        for candidato in candidatos
        if candidato
    }
    candidatos_tokens = [
        _column_tokens(candidato)
        for candidato in candidatos
        if candidato
    ]

    melhores: list[tuple[int, int, int, str]] = []
    for indice, cabecalho in enumerate(cabecalhos):
        normalizado = _normalize_column(cabecalho)
        tokens = _column_tokens(cabecalho)
        score = 0

        if normalizado in candidatos_normalizados:
            score = 100
        else:
            for tokens_candidato in candidatos_tokens:
                if not tokens_candidato:
                    continue
                intersecao = tokens & tokens_candidato
                # Aceita a mesma composicao mesmo se a ordem mudar:
                # "nome_fantasia_empresa" x "empresa.nome_fantasia".
                if tokens == tokens_candidato:
                    score = max(score, 92)
                elif tokens_candidato.issubset(tokens):
                    score = max(score, 82 + min(len(tokens_candidato), 8))
                elif tokens.issubset(tokens_candidato) and len(tokens) >= 2:
                    score = max(score, 72 + min(len(tokens), 8))
                elif len(intersecao) >= 2:
                    score = max(score, 55 + len(intersecao))

        # Regras especificas para o nome da empresa/unidade retornado pela MobileMed.
        if campo == "empresa.nome_fantasia":
            if "id" in tokens or "codigo" in tokens or "cnpj" in tokens:
                score -= 50
            if {"nome", "fantasia"}.issubset(tokens):
                score = max(score, 96)
            if {"empresa", "nome"}.issubset(tokens):
                score = max(score, 90)
            if {"unidade", "nome"}.issubset(tokens):
                score = max(score, 88)
            if {"company", "name"}.issubset(tokens):
                score = max(score, 88)
            if tokens in ({"empresa"}, {"unidade"}, {"fantasia"}):
                score = max(score, 70)

        if score > 0:
            preenchidos = _non_empty_count(cabecalho, linhas_csv)
            # Prioriza a coluna que realmente possui dados. O indice invertido
            # preserva a primeira ocorrencia em caso de empate.
            melhores.append((score, preenchidos, -indice, cabecalho))

    if melhores:
        melhores.sort(reverse=True)
        return melhores[0][3]

    sufixo = _normalize_column(campo.split(".")[-1])
    if len(sufixo) >= 4:
        possibilidades = []
        for indice, cabecalho in enumerate(cabecalhos):
            atual = _normalize_column(cabecalho)
            if atual.endswith(sufixo) or sufixo.endswith(atual):
                possibilidades.append(
                    (_non_empty_count(cabecalho, linhas_csv), -indice, cabecalho)
                )
        if possibilidades:
            possibilidades.sort(reverse=True)
            return possibilidades[0][2]
    return None


def _selected_fields(relatorio) -> list[str]:
    try:
        campos = json.loads(relatorio.campos or "[]")
    except (TypeError, json.JSONDecodeError):
        campos = []
    return [str(c) for c in campos if isinstance(c, str)]


def _project_rows(
    cabecalhos_csv: list[str],
    linhas_csv: list[dict[str, str]],
    campos_selecionados: list[str],
) -> tuple[list[str], list[dict[str, Any]]]:
    campos = campos_selecionados or list(cabecalhos_csv)
    mapeamento = {
        campo: _find_actual_header(campo, cabecalhos_csv, linhas_csv)
        for campo in campos
    }

    linhas: list[dict[str, Any]] = []
    for origem in linhas_csv:
        destino = {}
        for campo in campos:
            cabecalho = mapeamento.get(campo)
            destino[campo] = origem.get(cabecalho, "") if cabecalho else ""
        linhas.append(destino)
    return campos, linhas


def _safe_excel_value(valor: Any) -> Any:
    if valor is None:
        return ""
    if isinstance(valor, (int, float, bool, datetime, date)):
        return valor

    texto = str(valor)
    if texto.startswith(("=", "+", "@")):
        return "'" + texto
    if texto.startswith("-") and not re.fullmatch(r"-?\d+(?:[.,]\d+)?", texto.strip()):
        return "'" + texto
    return texto


def _build_workbook(
    colunas: list[str],
    linhas: list[dict[str, Any]],
    titulo: str,
    table_name: str,
    informacoes: dict[str, Any],
) -> io.BytesIO:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = titulo[:31]
    planilha.freeze_panes = "A2"
    planilha.append(colunas)

    preenchimento = PatternFill(fill_type="solid", fgColor="1F4E78")
    for celula in planilha[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal="center", vertical="center")

    larguras = {i: len(str(c)) for i, c in enumerate(colunas, start=1)}
    for linha in linhas:
        valores = [_safe_excel_value(linha.get(c, "")) for c in colunas]
        planilha.append(valores)
        for i, valor in enumerate(valores, start=1):
            larguras[i] = min(max(larguras.get(i, 10), len(str(valor))), 60)

    for i, largura in larguras.items():
        planilha.column_dimensions[get_column_letter(i)].width = max(12, largura + 2)

    if colunas:
        planilha.auto_filter.ref = planilha.dimensions
    if linhas and colunas:
        tabela = Table(displayName=table_name, ref=planilha.dimensions)
        tabela.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        planilha.add_table(tabela)

    info = workbook.create_sheet("Informacoes")
    info.append(["Informacao", "Valor"])
    for chave, valor in informacoes.items():
        info.append([chave, "" if valor is None else str(valor)])
    info.column_dimensions["A"].width = 30
    info.column_dimensions["B"].width = 80
    for celula in info[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = preenchimento

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_report_xlsx(relatorio) -> io.BytesIO:
    caminho = cache_report_csv(relatorio)
    cabecalhos, linhas_csv = _parse_csv(caminho.read_bytes())
    colunas, linhas = _project_rows(cabecalhos, linhas_csv, _selected_fields(relatorio))

    return _build_workbook(
        colunas=colunas,
        linhas=linhas,
        titulo="Dados",
        table_name="tb_mobilemed",
        informacoes={
            "Relatorio": relatorio.nome,
            "Relatorio ID": relatorio.id,
            "Request ID": relatorio.request_id,
            "Periodo inicial": relatorio.data_inicio,
            "Periodo final": relatorio.data_fim,
            "Solicitado por": relatorio.solicitado_por,
            "Solicitado em": relatorio.solicitado_em,
            "Concluido em": relatorio.concluido_em,
            "Quantidade de registros": len(linhas),
            "Gerado em": datetime.now().isoformat(),
        },
    )


def _row_key(linha: dict[str, Any]) -> str:
    partes: list[str] = []
    exame_id = str(linha.get("exame.id") or "").strip()
    if exame_id:
        partes.append(f"exame.id={exame_id}")
        for campo in (
            "usuario.id", "usuario.crm", "laudo_usuario.action",
            "subespecialidade.id",
        ):
            valor = str(linha.get(campo) or "").strip()
            if valor:
                partes.append(f"{campo}={valor}")
        return "|".join(partes)

    for campo in (
        "exame.pacs_accession_no", "exame.codigo_pedido", "exame.codigo_paciente",
    ):
        valor = str(linha.get(campo) or "").strip()
        if valor:
            return f"{campo}={valor}"

    serializado = json.dumps(linha, ensure_ascii=False, sort_keys=True, default=str)
    return "hash=" + hashlib.sha256(serializado.encode("utf-8")).hexdigest()


def regenerate_bi_xlsx() -> Path:
    from ..models.mobilemed_relatorio import MobilemedException as MobilemedRelatorio

    relatorios = (
        MobilemedRelatorio.query
        .filter_by(status="concluido")
        .order_by(MobilemedRelatorio.solicitado_em.asc())
        .all()
    )
    relatorios = [r for r in relatorios if is_bi_name(r.nome)]
    if not relatorios:
        raise ValueError("Nenhum relatorio BI concluido foi encontrado.")

    colunas_dados: list[str] = []
    lotes: list[tuple[Any, list[dict[str, Any]]]] = []
    erros: list[str] = []

    for relatorio in relatorios:
        try:
            caminho = cache_report_csv(relatorio)
            cabecalhos, linhas_csv = _parse_csv(caminho.read_bytes())
            campos = _selected_fields(relatorio)
            if "exame.id" not in campos:
                campos.insert(0, "exame.id")
            colunas, linhas = _project_rows(cabecalhos, linhas_csv, campos)
            for coluna in colunas:
                if coluna not in colunas_dados:
                    colunas_dados.append(coluna)
            lotes.append((relatorio, linhas))
        except Exception as erro:  # noqa: BLE001
            erros.append(f"Relatorio {relatorio.id}: {erro}")

    if not lotes:
        raise ValueError("Nenhum dos relatorios BI pode ser processado.")

    registros: dict[str, dict[str, Any]] = {}
    for relatorio, linhas in lotes:
        for linha in linhas:
            chave = _row_key(linha)
            atual = registros.get(chave, {})
            for coluna, valor in linha.items():
                if valor not in (None, ""):
                    atual[coluna] = valor
                elif coluna not in atual:
                    atual[coluna] = ""

            atual["_relatorio_id"] = relatorio.id
            atual["_relatorio_nome"] = relatorio.nome
            atual["_data_inicio"] = relatorio.data_inicio
            atual["_data_fim"] = relatorio.data_fim
            atual["_solicitado_em"] = (
                relatorio.solicitado_em.isoformat() if relatorio.solicitado_em else ""
            )
            atual["_concluido_em"] = (
                relatorio.concluido_em.isoformat() if relatorio.concluido_em else ""
            )
            registros[chave] = atual

    metadados = [
        "_relatorio_id", "_relatorio_nome", "_data_inicio", "_data_fim",
        "_solicitado_em", "_concluido_em",
    ]
    colunas = metadados + colunas_dados
    linhas = list(registros.values())

    buffer = _build_workbook(
        colunas=colunas,
        linhas=linhas,
        titulo="BI_Consolidado",
        table_name="tb_mobilemed_bi",
        informacoes={
            "Tipo": "Consolidado automatico MobileMed",
            "Relatorios processados": len(lotes),
            "Registros consolidados": len(linhas),
            "Relatorios ignorados por erro": len(erros),
            "Erros": " | ".join(erros),
            "Atualizado em": datetime.now().isoformat(),
        },
    )
    _atomic_write(BI_XLSX_PATH, buffer.getvalue())
    return BI_XLSX_PATH
