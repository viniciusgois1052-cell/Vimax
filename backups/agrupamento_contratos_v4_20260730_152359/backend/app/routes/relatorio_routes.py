from flask import Blueprint, jsonify, request, send_file
from sqlalchemy import func, extract, text, exists
from sqlalchemy.orm import joinedload, selectinload
from .. import db
from ..models.chamado import Chamado
from ..models.ativo import Ativo, ativo_contratos
from ..models.empresa import Empresa
from ..models.orcamento import Orcamento
from ..models.contrato import Contrato
from ..models.categoria_chamado import CategoriaChamado
from ..models.localizacao import Localizacao
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from io import BytesIO

relatorio_bp = Blueprint("relatorio_bp", __name__)

# ── helpers ──────────────────────────────────────────────────────────────────

def _style_header(ws, num_cols):
    """Aplica estilo de cabeçalho na primeira linha da planilha."""
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 20
    for i in range(1, num_cols + 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 22

def _make_excel(sheets):
    """
    sheets: lista de dicts { title, headers, rows }
    Retorna BytesIO com o arquivo Excel.
    """
    wb = Workbook()
    first = True
    for sheet in sheets:
        if first:
            ws = wb.active
            ws.title = sheet["title"][:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet["title"][:31])
        ws.append(sheet["headers"])
        for row in sheet["rows"]:
            ws.append(row)
        _style_header(ws, len(sheet["headers"]))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def _get_token(request):
    return request.headers.get("X-API-Token") or request.args.get("token")


def _ativo_sem_contrato():
    """Cláusula baseada na nova relação N:N entre ativos e contratos."""
    return ~exists().where(ativo_contratos.c.ativo_id == Ativo.id)

# ── DASHBOARD ────────────────────────────────────────────────────────────────

@relatorio_bp.route("/dashboard", methods=["GET"])
def dashboard_relatorios():
    try:
        # Filtros de data
        data_inicio_raw = (request.args.get('data_inicio') or '').strip()
        data_fim_raw = (request.args.get('data_fim') or '').strip()

        def parse_data_filtro(valor, campo):
            if not valor:
                return None
            try:
                return datetime.strptime(valor, '%Y-%m-%d')
            except ValueError:
                raise ValueError(
                    f"{campo} deve estar no formato YYYY-MM-DD"
                )

        try:
            data_inicio = parse_data_filtro(
                data_inicio_raw,
                'data_inicio'
            )
            data_fim = parse_data_filtro(
                data_fim_raw,
                'data_fim'
            )

            if data_fim:
                data_fim = data_fim.replace(
                    hour=23,
                    minute=59,
                    second=59,
                    microsecond=999999
                )

            if data_inicio and data_fim and data_inicio > data_fim:
                return jsonify({
                    'error': 'data_inicio não pode ser maior que data_fim'
                }), 400

        except ValueError as error:
            return jsonify({'error': str(error)}), 400

        def filtro_data_orm(query, campo):
            if data_inicio:
                query = query.filter(campo >= data_inicio)
            if data_fim:
                query = query.filter(campo <= data_fim)
            return query

        sql_clauses = []
        sql_params = {}

        if data_inicio:
            sql_clauses.append(
                'AND c.created_at >= :data_inicio'
            )
            sql_params['data_inicio'] = data_inicio

        if data_fim:
            sql_clauses.append(
                'AND c.created_at <= :data_fim'
            )
            sql_params['data_fim'] = data_fim

        filtro_sql = ' '.join(sql_clauses)

        # Estatísticas por Empresa
        empresas = Empresa.query.all()
        stats_map = {}
        for emp in empresas:
            stats_map[emp.id] = {
                "nome": emp.nome,
                "gasto_orcamentos": 0.0,
                "custo_chamados": 0.0,
                "ativos": 0,
                "ativos_sem_contrato": 0
            }

        for emp_id, total in (db.session.query(Orcamento.empresa_id, func.sum(Orcamento.valor))
                               .filter(Orcamento.status == 'Aprovado')
                               .group_by(Orcamento.empresa_id).all()):
            if emp_id in stats_map:
                stats_map[emp_id]["gasto_orcamentos"] = float(total or 0)

        q_custo = db.session.query(Chamado.empresa_id, func.sum(Chamado.valor_total)).filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
        q_custo = filtro_data_orm(q_custo, Chamado.created_at)
        for emp_id, total in q_custo.group_by(Chamado.empresa_id).all():
            if emp_id in stats_map:
                stats_map[emp_id]["custo_chamados"] = float(total or 0)

        for emp_id, qtd in db.session.query(Ativo.empresa_id, func.count(Ativo.id)).group_by(Ativo.empresa_id).all():
            if emp_id in stats_map:
                stats_map[emp_id]["ativos"] = int(qtd or 0)

        for emp_id, qtd in (db.session.query(Ativo.empresa_id, func.count(Ativo.id))
                             .filter(_ativo_sem_contrato())
                             .group_by(Ativo.empresa_id).all()):
            if emp_id in stats_map:
                stats_map[emp_id]["ativos_sem_contrato"] = int(qtd or 0)

        empresas_stats = sorted(list(stats_map.values()), key=lambda x: x["nome"])

        # Custos chamados por ativo (top 10)
        custos_ativos_chamados = []
        try:
            q = (db.session.query(Ativo.nome, func.sum(Chamado.valor_total))
                    .join(Chamado, Chamado.ativo_id == Ativo.id)
                    .filter(Chamado.ativo == True, Chamado.status != 'Cancelado'))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Ativo.nome).order_by(func.sum(Chamado.valor_total).desc()).limit(10).all()
            custos_ativos_chamados = [[r[0], float(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro custos_ativos_chamados: {e}")

        # Gastos orçamentos por ativo (top 10)
        gastos_ativos_orcamentos = []
        try:
            rows = (db.session.query(Ativo.nome, func.sum(Orcamento.valor))
                    .join(Orcamento, Ativo.orcamento_id == Orcamento.id)
                    .group_by(Ativo.nome)
                    .order_by(func.sum(Orcamento.valor).desc())
                    .limit(10).all())
            gastos_ativos_orcamentos = [[r[0], float(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro gastos_ativos_orcamentos: {e}")

        # Ativos sem contrato
        ativos_sem_contrato_count = (db.session.query(func.count(Ativo.id))
                                     .filter(_ativo_sem_contrato()).scalar() or 0)

        ativos_sem_contrato_detalhes = []
        try:
            rows = (db.session.query(Ativo.nome, Empresa.nome, Localizacao.nome)
                    .join(Empresa, Ativo.empresa_id == Empresa.id)
                    .outerjoin(Localizacao, Ativo.localizacao_id == Localizacao.id)
                    .filter(_ativo_sem_contrato())
                    .order_by(Empresa.nome, Ativo.nome).all())
            ativos_sem_contrato_detalhes = [[r[0], r[1], r[2] or 'Não informada'] for r in rows]
        except Exception as e:
            print(f"Erro ativos_sem_contrato_detalhes: {e}")

        # Chamados por categoria
        chamados_por_categoria = []
        try:
            q = (db.session.query(CategoriaChamado.nome, func.count(Chamado.id))
                    .join(Chamado, Chamado.categoria_id == CategoriaChamado.id)
                    .filter(Chamado.ativo == True, Chamado.status != 'Cancelado'))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(CategoriaChamado.nome).all()
            chamados_por_categoria = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_categoria: {e}")

        # Chamados por status
        chamados_por_status = []
        try:
            q = db.session.query(Chamado.status, func.count(Chamado.id)).filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Chamado.status).all()
            chamados_por_status = [[r[0] or 'Sem status', int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_status: {e}")

        # Chamados por tipo (maquinario / infraestrutura)
        chamados_por_tipo = []
        try:
            q = db.session.query(Chamado.tipo, func.count(Chamado.id)).filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Chamado.tipo).all()
            chamados_por_tipo = [[r[0] or 'Não informado', int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_tipo: {e}")

        # Chamados por prioridade
        chamados_por_prioridade = []
        try:
            q = db.session.query(Chamado.prioridade, func.count(Chamado.id)).filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Chamado.prioridade).all()
            chamados_por_prioridade = [[r[0] or 'Não informada', int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_prioridade: {e}")

        # Chamados por empresa
        chamados_por_empresa = []
        try:
            q = (db.session.query(Empresa.nome, func.count(Chamado.id))
                    .join(Chamado, Chamado.empresa_id == Empresa.id)
                    .filter(Chamado.ativo == True, Chamado.status != 'Cancelado'))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Empresa.nome).order_by(func.count(Chamado.id).desc()).all()
            chamados_por_empresa = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_empresa: {e}")

        # Chamados por maquinário (ativo)
        chamados_por_maquinario = []
        try:
            q = (db.session.query(Ativo.nome, func.count(Chamado.id))
                    .join(Chamado, Chamado.ativo_id == Ativo.id)
                    .filter(Chamado.ativo == True, Chamado.tipo == 'maquinario'))
            q = filtro_data_orm(q, Chamado.created_at)
            rows = q.group_by(Ativo.nome).order_by(func.count(Chamado.id).desc()).limit(15).all()
            chamados_por_maquinario = [[r[0], int(r[1] or 0)] for r in rows]
        except Exception as e:
            print(f"Erro chamados_por_maquinario: {e}")

                # Chamados por infraestrutura
        chamados_por_infraestrutura = []
        try:
            sql_infraestrutura = text(f"""
                SELECT
                    i.nome,
                    COUNT(c.id) AS total
                FROM chamados c
                JOIN infraestrutura i
                    ON c.infraestrutura_id = i.id
                WHERE c.ativo = 1
                  AND c.tipo = 'infraestrutura'
                  AND c.status != 'Cancelado'
                  {filtro_sql}
                GROUP BY i.nome
                ORDER BY total DESC
                LIMIT 15
            """)

            rows = db.session.execute(
                sql_infraestrutura,
                sql_params
            ).fetchall()

            chamados_por_infraestrutura = [
                [row[0], int(row[1] or 0)]
                for row in rows
            ]

        except Exception as e:
            print(f"Erro chamados_por_infraestrutura: {e}")

        # Tempo médio de solução (em dias)
        tempo_medio_solucao = []
        try:
            sql_tempo_solucao = text(f"""
                SELECT
                    e.nome,
                    ROUND(
                        AVG(
                            TIMESTAMPDIFF(
                                HOUR,
                                c.data_abertura,
                                c.data_solucao
                            ) / 24.0
                        ),
                        1
                    ) AS media_dias
                FROM chamados c
                JOIN empresas e
                    ON c.empresa_id = e.id
                WHERE c.ativo = 1
                  AND c.data_solucao IS NOT NULL
                  AND c.status != 'Cancelado'
                  {filtro_sql}
                GROUP BY e.nome
                ORDER BY media_dias DESC
            """)

            rows = db.session.execute(
                sql_tempo_solucao,
                sql_params
            ).fetchall()

            tempo_medio_solucao = [
                [row[0], float(row[1] or 0)]
                for row in rows
            ]

        except Exception as e:
            print(f"Erro tempo_medio_solucao: {e}")
            
        # Evolução mensal
        chamados_por_mes = []
        try:
            q_mes = (db.session.query(
                        extract('month', Chamado.created_at),
                        extract('year', Chamado.created_at),
                        func.count(Chamado.id))
                    .filter(Chamado.ativo == True, Chamado.status != 'Cancelado'))
            q_mes = filtro_data_orm(q_mes, Chamado.created_at)
            rows = (q_mes
                    .group_by(extract('year', Chamado.created_at), extract('month', Chamado.created_at))
                    .order_by(extract('year', Chamado.created_at).desc(), extract('month', Chamado.created_at).desc())
                    .limit(12).all())
            rows = list(reversed(rows))
            meses = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            for r in rows:
                try:
                    chamados_por_mes.append([f"{meses[int(r[0])]}/{str(int(r[1]))[-2:]}", int(r[2])])
                except:
                    continue
        except Exception as e:
            print(f"Erro chamados_por_mes: {e}")

        # Resumo geral
        total_ativos = db.session.query(func.count(Ativo.id)).scalar() or 0
        q_total_ch = db.session.query(func.count(Chamado.id)).filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
        q_total_ch = filtro_data_orm(q_total_ch, Chamado.created_at)
        total_chamados = q_total_ch.scalar() or 0
        total_gasto_orcamentos = (db.session.query(func.sum(Orcamento.valor))
                                  .filter(Orcamento.status == 'Aprovado').scalar() or 0)
        q_custo_total = db.session.query(func.sum(Chamado.valor_total)).filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
        q_custo_total = filtro_data_orm(q_custo_total, Chamado.created_at)
        total_custo_chamados = q_custo_total.scalar() or 0
        total_infraestruturas = db.session.execute(
            text("SELECT COUNT(*) FROM infraestrutura WHERE ativo = 1")).scalar() or 0

        # ── Custo por Fornecedor ──────────────────────────────────────────────────
        custo_por_fornecedor = []
        try:
            rows_forn = db.session.execute(text("""
                SELECT
                    f.id,
                    f.nome,
                    COUNT(o.id)                                          AS total_orcamentos,
                    COALESCE(SUM(o.valor), 0)                            AS total_orcado,
                    COALESCE(SUM(CASE WHEN o.status = 'Aprovado'  THEN o.valor ELSE 0 END), 0) AS aprovado,
                    COALESCE(SUM(CASE WHEN o.status = 'Pendente'  THEN o.valor ELSE 0 END), 0) AS pendente,
                    COALESCE(SUM(CASE WHEN o.status NOT IN ('Aprovado','Pendente') THEN o.valor ELSE 0 END), 0) AS outros
                FROM fornecedores f
                INNER JOIN orcamentos o ON o.fornecedor_id = f.id
                GROUP BY f.id, f.nome
                HAVING total_orcado > 0
                ORDER BY total_orcado DESC
            """)).fetchall()
            for r in rows_forn:
                custo_por_fornecedor.append({
                    'fornecedor_id':     r[0],
                    'fornecedor_nome':   r[1],
                    'total_orcamentos':  int(r[2] or 0),
                    'custo_total':       float(r[3] or 0),
                    'custo_aprovado':    float(r[4] or 0),
                    'custo_pendente':    float(r[5] or 0),
                    'custo_outros':      float(r[6] or 0),
                })
        except Exception as e:
            print(f"Erro custo_por_fornecedor: {e}")

        return jsonify({
            "custo_por_fornecedor": custo_por_fornecedor,
            "empresas_stats": empresas_stats,
            "custos_ativos_chamados": custos_ativos_chamados,
            "gastos_ativos_orcamentos": gastos_ativos_orcamentos,
            "ativos_sem_contrato": int(ativos_sem_contrato_count),
            "ativos_sem_contrato_detalhes": ativos_sem_contrato_detalhes,
            "chamados_por_categoria": chamados_por_categoria,
            "chamados_por_status": chamados_por_status,
            "chamados_por_tipo": chamados_por_tipo,
            "chamados_por_prioridade": chamados_por_prioridade,
            "chamados_por_empresa": chamados_por_empresa,
            "chamados_por_maquinario": chamados_por_maquinario,
            "chamados_por_infraestrutura": chamados_por_infraestrutura,
            "tempo_medio_solucao": tempo_medio_solucao,
            "chamados_por_mes": chamados_por_mes,
            "resumo": {
                "total_ativos": int(total_ativos),
                "total_chamados": int(total_chamados),
                "total_gasto_orcamentos": float(total_gasto_orcamentos or 0),
                "total_custo_chamados": float(total_custo_chamados or 0),
                "ativos_sem_contrato": int(ativos_sem_contrato_count),
                "total_infraestruturas": int(total_infraestruturas),
            }
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── EXPORTS ──────────────────────────────────────────────────────────────────

@relatorio_bp.route("/export/chamados_completo", methods=["GET"])
def export_chamados_completo():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.id,
                c.titulo,
                c.status,
                c.prioridade,
                c.tipo,
                c.data_abertura,
                c.data_solucao,
                ROUND(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0, 1) as dias_solucao,
                e.nome as empresa,
                cat.nome as categoria,
                a.nome as ativo,
                i.nome as infraestrutura,
                l.nome as localizacao,
                c.valor_total,
                c.descricao
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            LEFT JOIN ativos a ON c.ativo_id = a.id
            LEFT JOIN infraestrutura i ON c.infraestrutura_id = i.id
            LEFT JOIN localizacoes l ON c.localizacao_id = l.id
            WHERE c.ativo = 1
            ORDER BY c.data_abertura DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados Completo",
            "headers": ["ID", "Título", "Status", "Prioridade", "Tipo", "Data Abertura",
                        "Data Solução", "Dias p/ Solução", "Empresa", "Categoria",
                        "Maquinário", "Infraestrutura", "Localização", "Custo Total", "Descrição"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_completo.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_maquinario", methods=["GET"])
def export_chamados_por_maquinario():
    try:
        rows = db.session.execute(text("""
            SELECT
                a.nome as maquinario,
                e.nome as empresa,
                COUNT(c.id) as total_chamados,
                SUM(c.valor_total) as custo_total,
                AVG(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0) as media_dias_solucao
            FROM chamados c
            JOIN ativos a ON c.ativo_id = a.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.tipo = 'maquinario'
            GROUP BY a.nome, e.nome
            ORDER BY total_chamados DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Maquinário",
            "headers": ["Maquinário", "Empresa", "Total Chamados", "Custo Total", "Média Dias p/ Solução"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_maquinario.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_maquinario_categoria", methods=["GET"])
def export_chamados_por_maquinario_categoria():
    try:
        rows = db.session.execute(text("""
            SELECT
                a.nome as maquinario,
                cat.nome as categoria,
                e.nome as empresa,
                COUNT(c.id) as total_chamados,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            JOIN ativos a ON c.ativo_id = a.id
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.tipo = 'maquinario'
            GROUP BY a.nome, cat.nome, e.nome
            ORDER BY a.nome, total_chamados DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Maquinário x Categoria",
            "headers": ["Maquinário", "Categoria", "Empresa", "Total Chamados", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_maquinario_categoria.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_infraestrutura", methods=["GET"])
def export_chamados_por_infraestrutura():
    try:
        rows = db.session.execute(text("""
            SELECT
                i.nome as infraestrutura,
                e.nome as empresa,
                COUNT(c.id) as total_chamados,
                SUM(c.valor_total) as custo_total,
                AVG(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0) as media_dias_solucao
            FROM chamados c
            JOIN infraestrutura i ON c.infraestrutura_id = i.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.tipo = 'infraestrutura' AND c.status != 'Cancelado'
            GROUP BY i.nome, e.nome
            ORDER BY total_chamados DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Infraestrutura",
            "headers": ["Infraestrutura", "Empresa", "Total Chamados", "Custo Total", "Média Dias p/ Solução"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_infraestrutura.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_empresa", methods=["GET"])
def export_chamados_por_empresa():
    try:
        rows = db.session.execute(text("""
            SELECT
                e.nome as empresa,
                c.status,
                c.prioridade,
                c.tipo,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY e.nome, c.status, c.prioridade, c.tipo
            ORDER BY e.nome, total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Empresa",
            "headers": ["Empresa", "Status", "Prioridade", "Tipo", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_empresa.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_categoria", methods=["GET"])
def export_chamados_por_categoria():
    try:
        rows = db.session.execute(text("""
            SELECT
                cat.nome as categoria,
                e.nome as empresa,
                c.tipo,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY cat.nome, e.nome, c.tipo
            ORDER BY total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Categoria",
            "headers": ["Categoria", "Empresa", "Tipo", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_categoria.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_status", methods=["GET"])
def export_chamados_por_status():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.status,
                e.nome as empresa,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY c.status, e.nome
            ORDER BY total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Status",
            "headers": ["Status", "Empresa", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_status.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_prioridade", methods=["GET"])
def export_chamados_por_prioridade():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.prioridade,
                e.nome as empresa,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1
            GROUP BY c.prioridade, e.nome
            ORDER BY total DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Chamados por Prioridade",
            "headers": ["Prioridade", "Empresa", "Total", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_prioridade.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/chamados_por_mes", methods=["GET"])
def export_chamados_por_mes():
    try:
        rows = db.session.execute(text("""
            SELECT
                DATE_FORMAT(c.created_at, '%m/%Y') as mes_ano,
                COUNT(c.id) as total,
                SUM(c.valor_total) as custo_total
            FROM chamados c
            WHERE c.ativo = 1
            GROUP BY DATE_FORMAT(c.created_at, '%m/%Y'), DATE_FORMAT(c.created_at, '%Y-%m')
            ORDER BY DATE_FORMAT(c.created_at, '%Y-%m') DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Evolução Mensal",
            "headers": ["Mês/Ano", "Total Chamados", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="chamados_por_mes.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/tempo_solucao", methods=["GET"])
def export_tempo_solucao():
    try:
        rows = db.session.execute(text("""
            SELECT
                c.id,
                c.titulo,
                e.nome as empresa,
                c.status,
                c.prioridade,
                c.tipo,
                c.data_abertura,
                c.data_solucao,
                ROUND(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao) / 24.0, 1) as dias_solucao
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            WHERE c.ativo = 1 AND c.data_solucao IS NOT NULL AND c.status != 'Cancelado'
            ORDER BY dias_solucao DESC
        """)).fetchall()

        buf = _make_excel([{
            "title": "Tempo de Solução",
            "headers": ["ID", "Título", "Empresa", "Status", "Prioridade", "Tipo",
                        "Data Abertura", "Data Solução", "Dias p/ Solução"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="tempo_solucao.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/custos_ativos_chamados", methods=["GET"])
def export_custos_ativos_chamados():
    try:
        rows = (db.session.query(Ativo.nome, Empresa.nome, func.count(Chamado.id), func.sum(Chamado.valor_total))
                .join(Chamado, Chamado.ativo_id == Ativo.id)
                .join(Empresa, Ativo.empresa_id == Empresa.id)
                .filter(Chamado.ativo == True, Chamado.status != 'Cancelado')
                .group_by(Ativo.nome, Empresa.nome)
                .order_by(func.sum(Chamado.valor_total).desc()).all())

        buf = _make_excel([{
            "title": "Custos Chamados por Ativo",
            "headers": ["Ativo", "Empresa", "Total Chamados", "Custo Total"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="custos_ativos_chamados.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/gastos_ativos_orcamentos", methods=["GET"])
def export_gastos_ativos_orcamentos():
    try:
        rows = (db.session.query(Ativo.nome, func.sum(Orcamento.valor))
                .join(Orcamento, Ativo.orcamento_id == Orcamento.id)
                .group_by(Ativo.nome)
                .order_by(func.sum(Orcamento.valor).desc()).all())

        buf = _make_excel([{
            "title": "Gastos Orçamentos por Ativo",
            "headers": ["Ativo", "Gasto Total Orçamentos"],
            "rows": [list(r) for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="gastos_ativos_orcamentos.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/ativos_sem_contrato", methods=["GET"])
def export_ativos_sem_contrato():
    try:
        rows = (db.session.query(Ativo.nome, Empresa.nome, Localizacao.nome)
                .join(Empresa, Ativo.empresa_id == Empresa.id)
                .outerjoin(Localizacao, Ativo.localizacao_id == Localizacao.id)
                .filter(_ativo_sem_contrato())
                .order_by(Empresa.nome, Ativo.nome).all())

        buf = _make_excel([{
            "title": "Ativos sem Contrato",
            "headers": ["Ativo", "Empresa", "Localização"],
            "rows": [[r[0], r[1], r[2] or 'Não informada'] for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="ativos_sem_contrato.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/custo_por_fornecedor", methods=["GET"])
def export_custo_por_fornecedor():
    try:
        rows = db.session.execute(text("""
            SELECT
                f.nome,
                COUNT(o.id),
                COALESCE(SUM(o.valor), 0),
                COALESCE(SUM(CASE WHEN o.status = 'Aprovado' THEN o.valor ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN o.status = 'Pendente' THEN o.valor ELSE 0 END), 0)
            FROM fornecedores f
            INNER JOIN orcamentos o ON o.fornecedor_id = f.id
            GROUP BY f.id, f.nome
            HAVING COALESCE(SUM(o.valor), 0) > 0
            ORDER BY SUM(o.valor) DESC        """)).fetchall()

        buf = _make_excel([{
            "title": "Custo por Fornecedor",
            "headers": ["Fornecedor", "Qtd Orçamentos", "Total Orçado (R$)", "Aprovado (R$)", "Pendente (R$)"],
            "rows": [[r[0], int(r[1]), float(r[2]), float(r[3]), float(r[4])] for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="custo_por_fornecedor.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/empresas_stats", methods=["GET"])
def export_empresas_stats():
    try:
        empresas = Empresa.query.all()
        stats_map = {}
        for emp in empresas:
            stats_map[emp.id] = {"nome": emp.nome, "gasto_orcamentos": 0.0, "custo_chamados": 0.0, "ativos": 0, "ativos_sem_contrato": 0}

        for emp_id, total in (db.session.query(Orcamento.empresa_id, func.sum(Orcamento.valor))
                               .filter(Orcamento.status == 'Aprovado').group_by(Orcamento.empresa_id).all()):
            if emp_id in stats_map: stats_map[emp_id]["gasto_orcamentos"] = float(total or 0)

        for emp_id, total in (db.session.query(Chamado.empresa_id, func.sum(Chamado.valor_total))
                               .filter(Chamado.ativo == True, Chamado.status != 'Cancelado').group_by(Chamado.empresa_id).all()):
            if emp_id in stats_map: stats_map[emp_id]["custo_chamados"] = float(total or 0)

        for emp_id, qtd in db.session.query(Ativo.empresa_id, func.count(Ativo.id)).group_by(Ativo.empresa_id).all():
            if emp_id in stats_map: stats_map[emp_id]["ativos"] = int(qtd or 0)

        for emp_id, qtd in (db.session.query(Ativo.empresa_id, func.count(Ativo.id))
                             .filter(_ativo_sem_contrato()).group_by(Ativo.empresa_id).all()):
            if emp_id in stats_map: stats_map[emp_id]["ativos_sem_contrato"] = int(qtd or 0)

        rows = sorted(list(stats_map.values()), key=lambda x: x["nome"])
        buf = _make_excel([{
            "title": "Estatísticas por Empresa",
            "headers": ["Empresa", "Qtd. Ativos", "Ativos sem Contrato", "Custo Chamados", "Gasto Orçamentos"],
            "rows": [[r["nome"], r["ativos"], r["ativos_sem_contrato"], r["custo_chamados"], r["gasto_orcamentos"]] for r in rows]
        }])
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="empresas_stats.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@relatorio_bp.route("/export/tudo", methods=["GET"])
def export_tudo():
    """Exporta todas as planilhas em um único arquivo Excel com múltiplas abas."""
    try:
        sheets = []

        # Aba 1 - Chamados Completo
        rows1 = db.session.execute(text("""
            SELECT c.id, c.titulo, c.status, c.prioridade, c.tipo,
                   c.data_abertura, c.data_solucao,
                   ROUND(TIMESTAMPDIFF(HOUR, c.data_abertura, c.data_solucao)/24.0,1),
                   e.nome, cat.nome, a.nome, i.nome, l.nome, c.valor_total
            FROM chamados c
            LEFT JOIN empresas e ON c.empresa_id = e.id
            LEFT JOIN categorias_chamado cat ON c.categoria_id = cat.id
            LEFT JOIN ativos a ON c.ativo_id = a.id
            LEFT JOIN infraestrutura i ON c.infraestrutura_id = i.id
            LEFT JOIN localizacoes l ON c.localizacao_id = l.id
            WHERE c.ativo = 1 ORDER BY c.data_abertura DESC
        """)).fetchall()
        sheets.append({"title": "Chamados Completo",
                        "headers": ["ID","Título","Status","Prioridade","Tipo","Abertura","Solução","Dias","Empresa","Categoria","Maquinário","Infraestrutura","Localização","Custo"],
                        "rows": [list(r) for r in rows1]})

        # Aba 2 - Por Maquinário
        rows2 = db.session.execute(text("""
            SELECT a.nome, e.nome, COUNT(c.id), SUM(c.valor_total),
                   ROUND(AVG(TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao)/24.0),1)
            FROM chamados c JOIN ativos a ON c.ativo_id=a.id
            JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 AND c.tipo='maquinario'
            GROUP BY a.nome,e.nome ORDER BY COUNT(c.id) DESC
        """)).fetchall()
        sheets.append({"title": "Por Maquinário",
                        "headers": ["Maquinário","Empresa","Total","Custo","Média Dias"],
                        "rows": [list(r) for r in rows2]})

        # Aba 3 - Por Infraestrutura
        rows3 = db.session.execute(text("""
            SELECT i.nome, e.nome, COUNT(c.id), SUM(c.valor_total),
                   ROUND(AVG(TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao)/24.0),1)
            FROM chamados c JOIN infraestrutura i ON c.infraestrutura_id=i.id
            JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 AND c.tipo='infraestrutura'
            GROUP BY i.nome,e.nome ORDER BY COUNT(c.id) DESC
        """)).fetchall()
        sheets.append({"title": "Por Infraestrutura",
                        "headers": ["Infraestrutura","Empresa","Total","Custo","Média Dias"],
                        "rows": [list(r) for r in rows3]})

        # Aba 4 - Por Empresa
        rows4 = db.session.execute(text("""
            SELECT e.nome, c.status, c.prioridade, c.tipo, COUNT(c.id), SUM(c.valor_total)
            FROM chamados c JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 GROUP BY e.nome,c.status,c.prioridade,c.tipo ORDER BY e.nome
        """)).fetchall()
        sheets.append({"title": "Por Empresa",
                        "headers": ["Empresa","Status","Prioridade","Tipo","Total","Custo"],
                        "rows": [list(r) for r in rows4]})

        # Aba 5 - Por Categoria
        rows5 = db.session.execute(text("""
            SELECT cat.nome, e.nome, c.tipo, COUNT(c.id), SUM(c.valor_total)
            FROM chamados c LEFT JOIN categorias_chamado cat ON c.categoria_id=cat.id
            JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 GROUP BY cat.nome,e.nome,c.tipo ORDER BY COUNT(c.id) DESC
        """)).fetchall()
        sheets.append({"title": "Por Categoria",
                        "headers": ["Categoria","Empresa","Tipo","Total","Custo"],
                        "rows": [list(r) for r in rows5]})

        # Aba 6 - Evolução Mensal
        rows6 = db.session.execute(text("""
            SELECT DATE_FORMAT(created_at,'%m/%Y'), COUNT(id), SUM(valor_total)
            FROM chamados WHERE ativo=1
            GROUP BY DATE_FORMAT(created_at,'%m/%Y'), DATE_FORMAT(created_at,'%Y-%m')
            ORDER BY DATE_FORMAT(created_at,'%Y-%m')
        """)).fetchall()
        sheets.append({"title": "Evolução Mensal",
                        "headers": ["Mês/Ano","Total","Custo Total"],
                        "rows": [list(r) for r in rows6]})

        # Aba 7 - Tempo de Solução
        rows7 = db.session.execute(text("""
            SELECT c.id, c.titulo, e.nome, c.prioridade, c.tipo,
                   c.data_abertura, c.data_solucao,
                   ROUND(TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao)/24.0,1)
            FROM chamados c LEFT JOIN empresas e ON c.empresa_id=e.id
            WHERE c.ativo=1 AND c.data_solucao IS NOT NULL
            ORDER BY TIMESTAMPDIFF(HOUR,c.data_abertura,c.data_solucao) DESC
        """)).fetchall()
        sheets.append({"title": "Tempo de Solução",
                        "headers": ["ID","Título","Empresa","Prioridade","Tipo","Abertura","Solução","Dias"],
                        "rows": [list(r) for r in rows7]})

        # Aba 8 - Ativos sem Contrato
        rows8 = (db.session.query(Ativo.nome, Empresa.nome, Localizacao.nome)
                 .join(Empresa, Ativo.empresa_id == Empresa.id)
                 .outerjoin(Localizacao, Ativo.localizacao_id == Localizacao.id)
                 .filter(_ativo_sem_contrato())
                 .order_by(Empresa.nome, Ativo.nome).all())
        sheets.append({"title": "Ativos sem Contrato",
                        "headers": ["Ativo","Empresa","Localização"],
                        "rows": [[r[0], r[1], r[2] or 'Não informada'] for r in rows8]})

        buf = _make_excel(sheets)
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"relatorio_completo_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ── CUSTOS DE CONTRATOS (para Relatórios) ────────────────────────────────────
@relatorio_bp.route("/custos_contratos", methods=["GET"])
def relatorio_custos_contratos():
    import urllib.request, json as _json
    from datetime import date
    from ..models.contrato import Contrato
    from ..utils.filters import apply_entity_filter

    # Cotação USD
    cotacao_usd = 5.0
    try:
        with urllib.request.urlopen('https://economia.awesomeapi.com.br/json/last/USD-BRL', timeout=5) as r:
            cotacao_usd = float(_json.loads(r.read())['USDBRL']['bid'])
    except Exception:
        pass

    hoje          = date.today()
    mes_atual     = hoje.month
    ano_atual     = hoje.year

    contratos = (
        Contrato.query
        .options(
            joinedload(Contrato.empresa),
            joinedload(Contrato.fornecedor),
            selectinload(Contrato.ativos_vinculados)
                .selectinload(Ativo.fornecedores)
        )
        .all()
    )

    # ── agrupadores ──────────────────────────────────────────────────────────
    por_empresa    = {}   # empresa_id -> {...}
    por_maquinario = {}   # ativo_nome -> {...}
    total_mensal   = 0.0
    total_anual    = 0.0

    for c in contratos:
        moeda = (c.moeda or 'BRL').upper()

        # valor mensal em BRL
        if c.is_mensal:
            v_mensal = c.valor
        else:
            if c.data_inicio and c.data_fim:
                meses = max(1, (c.data_fim.year - c.data_inicio.year) * 12 +
                               (c.data_fim.month - c.data_inicio.month))
                v_mensal = c.valor / meses
            else:
                v_mensal = c.valor

        if moeda == 'USD':
            v_mensal = v_mensal * cotacao_usd

        v_anual = v_mensal * mes_atual

        empresa_id   = c.empresa_id or 0
        empresa_nome = c.empresa.nome if c.empresa else 'Sem Empresa'

        # ── por empresa ──────────────────────────────────────────────────────
        if empresa_id not in por_empresa:
            por_empresa[empresa_id] = {
                'empresa_id':   empresa_id,
                'empresa_nome': empresa_nome,
                'total_mensal': 0.0,
                'total_anual':  0.0,
                'contratos':    []
            }
        por_empresa[empresa_id]['total_mensal'] += v_mensal
        por_empresa[empresa_id]['total_anual']  += v_anual
        por_empresa[empresa_id]['contratos'].append({
            'contrato_id':      c.id,
            'contrato_numero':  c.numero,
            'fornecedor_id':    c.fornecedor_id,
            'fornecedor_nome':  c.fornecedor.nome if c.fornecedor else 'Não informado',
            'ativos_quantidade': len(c.ativos_vinculados or []),
            'moeda':            moeda,
            'valor_original':   c.valor,
            'valor_mensal_brl': round(v_mensal, 2),
            'valor_anual_brl':  round(v_anual, 2),
            'is_mensal':        c.is_mensal,
        })

        # ── por maquinário (relação N:N) ───────────────────────────────────
        ativos = list(c.ativos_vinculados or [])
        if ativos:
            for ativo in ativos:
                key = f'ativo-{ativo.id}'
                fornecedores_nomes = {
                    fornecedor.nome
                    for fornecedor in (ativo.fornecedores or [])
                    if fornecedor.nome
                }
                if c.fornecedor and c.fornecedor.nome:
                    fornecedores_nomes.add(c.fornecedor.nome)

                if key not in por_maquinario:
                    por_maquinario[key] = {
                        'chave':               key,
                        'maquinario_id':       ativo.id,
                        'maquinario':          ativo.nome or '—',
                        'numero_serie':        ativo.numero_serie,
                        'empresa_nome':        (
                            ativo.empresa.nome
                            if ativo.empresa
                            else empresa_nome
                        ),
                        'fornecedores_nomes':  sorted(fornecedores_nomes),
                        'total_mensal':        0.0,
                        'total_anual':         0.0,
                        'contratos':           []
                    }
                else:
                    atuais = set(
                        por_maquinario[key].get('fornecedores_nomes') or []
                    )
                    por_maquinario[key]['fornecedores_nomes'] = sorted(
                        atuais | fornecedores_nomes
                    )

                por_maquinario[key]['total_mensal'] += v_mensal
                por_maquinario[key]['total_anual'] += v_anual
                por_maquinario[key]['contratos'].append({
                    'contrato_id':      c.id,
                    'contrato_numero':  c.numero,
                    'fornecedor_id':    c.fornecedor_id,
                    'fornecedor_nome':  (
                        c.fornecedor.nome
                        if c.fornecedor
                        else 'Não informado'
                    ),
                    'empresa_nome':     empresa_nome,
                    'moeda':            moeda,
                    'valor_original':   c.valor,
                    'valor_mensal_brl': round(v_mensal, 2),
                    'valor_anual_brl':  round(v_anual, 2),
                })
        else:
            key = 'sem-ativo'
            if key not in por_maquinario:
                por_maquinario[key] = {
                    'chave':              key,
                    'maquinario_id':      None,
                    'maquinario':         'Sem Ativo Vinculado',
                    'numero_serie':       None,
                    'empresa_nome':       empresa_nome,
                    'fornecedores_nomes': [],
                    'total_mensal':       0.0,
                    'total_anual':        0.0,
                    'contratos':          []
                }

            por_maquinario[key]['total_mensal'] += v_mensal
            por_maquinario[key]['total_anual'] += v_anual
            por_maquinario[key]['contratos'].append({
                'contrato_id':      c.id,
                'contrato_numero':  c.numero,
                'fornecedor_id':    c.fornecedor_id,
                'fornecedor_nome':  (
                    c.fornecedor.nome
                    if c.fornecedor
                    else 'Não informado'
                ),
                'empresa_nome':     empresa_nome,
                'moeda':            moeda,
                'valor_original':   c.valor,
                'valor_mensal_brl': round(v_mensal, 2),
                'valor_anual_brl':  round(v_anual, 2),
            })

        total_mensal += v_mensal
        total_anual  += v_anual

    # arredonda e converte para listas ordenadas
    lista_empresas = sorted(por_empresa.values(), key=lambda x: x['empresa_nome'])
    for e in lista_empresas:
        e['total_mensal'] = round(e['total_mensal'], 2)
        e['total_anual']  = round(e['total_anual'],  2)

    lista_maquinarios = sorted(por_maquinario.values(), key=lambda x: -x['total_mensal'])
    for m in lista_maquinarios:
        m['total_mensal'] = round(m['total_mensal'], 2)
        m['total_anual']  = round(m['total_anual'],  2)

    return jsonify({
        'cotacao_usd':        round(cotacao_usd, 4),
        'mes_referencia':     f"{mes_atual:02d}/{ano_atual}",
        'meses_acumulados':   mes_atual,
        'total_mensal':       round(total_mensal, 2),
        'total_anual':        round(total_anual,  2),
        'por_empresa':        lista_empresas,
        'por_maquinario':     lista_maquinarios,
    }), 200


# ── EXPORT: Custos de Contratos ───────────────────────────────────────────────
@relatorio_bp.route("/export/custos_contratos", methods=["GET"])
def export_custos_contratos():
    import urllib.request, json as _json
    from datetime import date
    from ..models.contrato import Contrato

    cotacao_usd = 5.0
    try:
        with urllib.request.urlopen('https://economia.awesomeapi.com.br/json/last/USD-BRL', timeout=5) as r:
            cotacao_usd = float(_json.loads(r.read())['USDBRL']['bid'])
    except Exception:
        pass

    hoje      = date.today()
    mes_atual = hoje.month

    contratos = (
        Contrato.query
        .options(
            joinedload(Contrato.empresa),
            joinedload(Contrato.fornecedor),
            selectinload(Contrato.ativos_vinculados)
                .selectinload(Ativo.fornecedores)
        )
        .all()
    )

    rows_emp = []
    rows_maq = []

    for c in contratos:
        moeda    = (c.moeda or 'BRL').upper()
        if c.is_mensal:
            v_mensal = c.valor
        else:
            if c.data_inicio and c.data_fim:
                meses = max(1, (c.data_fim.year - c.data_inicio.year) * 12 +
                               (c.data_fim.month - c.data_inicio.month))
                v_mensal = c.valor / meses
            else:
                v_mensal = c.valor
        if moeda == 'USD':
            v_mensal = v_mensal * cotacao_usd
        v_anual = round(v_mensal * mes_atual, 2)
        v_mensal = round(v_mensal, 2)

        empresa_nome = c.empresa.nome if c.empresa else 'Sem Empresa'
        fornecedor_contrato = (
            c.fornecedor.nome if c.fornecedor else 'Não informado'
        )
        nomes_ativos = ', '.join(
            sorted({ativo.nome for ativo in (c.ativos_vinculados or []) if ativo.nome})
        ) or 'Sem Ativo'

        rows_emp.append([
            empresa_nome,
            c.numero,
            fornecedor_contrato,
            len(c.ativos_vinculados or []),
            nomes_ativos,
            moeda,
            c.valor,
            v_mensal,
            v_anual
        ])

        ativos = list(c.ativos_vinculados or [])
        if ativos:
            for ativo in ativos:
                fornecedores_ativo = ', '.join(
                    sorted({
                        fornecedor.nome
                        for fornecedor in (ativo.fornecedores or [])
                        if fornecedor.nome
                    })
                ) or 'Não informado'
                rows_maq.append([
                    ativo.id,
                    ativo.nome,
                    ativo.numero_serie or '',
                    ativo.empresa.nome if ativo.empresa else empresa_nome,
                    fornecedores_ativo,
                    c.numero,
                    fornecedor_contrato,
                    moeda,
                    c.valor,
                    v_mensal,
                    v_anual
                ])
        else:
            rows_maq.append([
                '',
                'Sem Ativo',
                '',
                empresa_nome,
                'Não informado',
                c.numero,
                fornecedor_contrato,
                moeda,
                c.valor,
                v_mensal,
                v_anual
            ])

    buf = _make_excel([
        {
            'title': 'Custo por Empresa',
            'headers': [
                'Empresa', 'Contrato', 'Fornecedor do Contrato',
                'Qtd. Ativos', 'Ativos', 'Moeda', 'Valor Original',
                'Mensal (BRL)', 'Acumulado (BRL)'
            ],
            'rows': rows_emp
        },
        {
            'title': 'Custo por Maquinário',
            'headers': [
                'ID do Ativo', 'Maquinário', 'Número de Série', 'Empresa',
                'Fornecedores do Ativo', 'Contrato', 'Fornecedor do Contrato',
                'Moeda', 'Valor Original', 'Mensal (BRL)', 'Acumulado (BRL)'
            ],
            'rows': rows_maq
        }
    ])
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'custos_contratos_{hoje.strftime("%Y%m%d")}.xlsx')